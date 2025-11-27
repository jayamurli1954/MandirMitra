"""
Donation API Endpoints
"""

from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse, Response
from sqlalchemy.orm import Session
from sqlalchemy import func, and_
from typing import List, Optional
from datetime import datetime, date, timedelta
import io
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as ExcelImage
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
import os
import requests
from urllib.parse import urlparse

from app.core.database import get_db
from app.core.security import get_current_user
from app.core.audit import log_action, get_entity_dict
from fastapi import Request
from app.models.donation import Donation, DonationCategory, DonationType, InKindDonationSubType
from app.models.devotee import Devotee
from app.models.temple import Temple
from app.models.user import User
from app.models.accounting import Account, JournalEntry, JournalLine, JournalEntryStatus, TransactionType
from pydantic import BaseModel

router = APIRouter(prefix="/api/v1/donations", tags=["donations"])


@router.get("/categories/", response_model=List[dict])
def get_donation_categories(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Get list of donation categories
    """
    categories = db.query(DonationCategory).filter(
        DonationCategory.is_active == True
    ).all()
    
    return [
        {
            "id": cat.id,
            "name": cat.name,
            "description": cat.description,
            "is_80g_eligible": cat.is_80g_eligible,
            "account_id": cat.account_id
        }
        for cat in categories
    ]


def post_donation_to_accounting(db: Session, donation: Donation, temple_id: int):
    """
    Create journal entry for donation in accounting system
    
    For Cash Donations:
    Dr: Cash/Bank Account (based on payment mode)
    Cr: Donation Income Account
    
    For In-Kind Donations:
    - Inventory: Dr. Inventory Asset (1300), Cr. Donation Income
    - Assets: Dr. Asset Account (based on asset type), Cr. Donation Income
    - Event Sponsorship: Dr. Prepaid Expense or Expense, Cr. Donation Income

    Behaviour:
    - If donation category is linked to a specific income account -> credit that account
    - Otherwise -> credit 4100 (Donation Income - Main)

    This keeps Trial Balance short (single donation income line),
    while still allowing detailed category-wise accounts for temples that link them.
    """
    try:
        # Handle in-kind donations differently
        if donation.donation_type == DonationType.IN_KIND:
            # Determine debit account based on in-kind subtype
            debit_account_code = None
            if donation.inkind_subtype == InKindDonationSubType.INVENTORY:
                debit_account_code = '1300'  # Inventory Asset
            elif donation.inkind_subtype == InKindDonationSubType.ASSET:
                # Determine asset account based on asset type
                if donation.purity:  # Precious items (gold, silver)
                    debit_account_code = '1500'  # Precious Assets (or specific account)
                else:
                    debit_account_code = '1400'  # Fixed Assets
            elif donation.inkind_subtype == InKindDonationSubType.EVENT_SPONSORSHIP:
                # For event sponsorship, debit prepaid expense or expense account
                debit_account_code = '5100'  # Prepaid Expenses (or expense when event occurs)
            else:
                debit_account_code = '1300'  # Default to inventory
            
            # Get debit account
            debit_account = db.query(Account).filter(
                Account.temple_id == temple_id,
                Account.account_code == debit_account_code
            ).first()
            
            if not debit_account:
                # Try to find by account_subtype
                from app.models.accounting import AccountSubType
                if donation.inkind_subtype == InKindDonationSubType.INVENTORY:
                    debit_account = db.query(Account).filter(
                        Account.temple_id == temple_id,
                        Account.account_subtype == AccountSubType.INVENTORY
                    ).first()
                elif donation.inkind_subtype == InKindDonationSubType.ASSET:
                    debit_account = db.query(Account).filter(
                        Account.temple_id == temple_id,
                        Account.account_subtype == AccountSubType.PRECIOUS_ASSET
                    ).first()
        else:
            # Cash donation - determine debit account (payment method)
            debit_account_code = None
            if donation.payment_mode and donation.payment_mode.upper() in ['CASH', 'COUNTER']:
                debit_account_code = '1101'  # Cash in Hand - Counter
            elif donation.payment_mode and donation.payment_mode.upper() in ['UPI', 'ONLINE', 'CARD', 'NETBANKING']:
                debit_account_code = '1110'  # Bank - SBI Current Account
            elif donation.payment_mode and 'HUNDI' in donation.payment_mode.upper():
                debit_account_code = '1102'  # Cash in Hand - Hundi
            else:
                debit_account_code = '1101'  # Default to cash counter

            # Get debit account
            debit_account = db.query(Account).filter(
                Account.temple_id == temple_id,
                Account.account_code == debit_account_code
            ).first()

        # Determine credit account - PRIORITY: Category-linked account
        credit_account = None

        # First, try to use category-linked account (PRIORITY)
        if donation.category and hasattr(donation.category, 'account_id') and donation.category.account_id:
            credit_account = db.query(Account).filter(Account.id == donation.category.account_id).first()
            if credit_account:
                print(f"  Using category-linked account: {credit_account.account_code} - {credit_account.account_name}")

        # Fallback:
        # Use 4100 - Donation Income (Main) as default when category is not linked.
        # DO NOT use payment-mode accounts (4102, 4103) - all should go to income accounts.
        if not credit_account:
            credit_account_code = '4100'  # Donation Income - Main (parent)
            credit_account = db.query(Account).filter(
                Account.temple_id == temple_id,
                Account.account_code == credit_account_code
            ).first()

            if credit_account:
                print(f"  INFO: Category '{donation.category.name if donation.category else 'Unknown'}' not linked to account. Using main donation income account: {credit_account_code}")
            else:
                print(f"  ERROR: Main donation income account {credit_account_code} not found. Please ensure it exists in Chart of Accounts or link category to a specific account.")

        if not debit_account:
            error_msg = f"Debit account ({debit_account_code}) not found for temple {temple_id}. Please create the account in Chart of Accounts."
            print(f"ERROR: {error_msg}")
            print(f"  - Donation: {donation.receipt_number}")
            print(f"  - Payment mode: {donation.payment_mode}")
            return None
        
        if not credit_account:
            error_msg = f"Credit account not found for donation category '{donation.category.name if donation.category else 'Unknown'}'. Please link an account to the donation category or create default income accounts."
            print(f"ERROR: {error_msg}")
            print(f"  - Donation: {donation.receipt_number}")
            print(f"  - Category: {donation.category.name if donation.category else 'None'}")
            print(f"  - Category account_id: {donation.category.account_id if donation.category and hasattr(donation.category, 'account_id') else 'NO'}")
            return None

        # Create narration
        if donation.donation_type == DonationType.IN_KIND:
            item_desc = donation.item_name or "In-kind donation"
            narration = f"In-kind donation from {donation.devotee.name if donation.devotee else 'Anonymous'}: {item_desc}"
            if donation.category:
                narration += f" - {donation.category.name}"
        else:
            narration = f"Donation from {donation.devotee.name if donation.devotee else 'Anonymous'}"
            if donation.category:
                narration += f" - {donation.category.name}"

        # Generate entry number first
        year = donation.donation_date.year
        prefix = f"JE/{year}/"
        
        # Get last entry number for this year
        last_entry = db.query(JournalEntry).filter(
            JournalEntry.temple_id == temple_id,
            JournalEntry.entry_number.like(f"{prefix}%")
        ).order_by(JournalEntry.id.desc()).first()

        if last_entry:
            # Extract number and increment
            try:
                last_num = int(last_entry.entry_number.split('/')[-1])
                new_num = last_num + 1
            except:
                new_num = 1
        else:
            new_num = 1

        entry_number = f"{prefix}{new_num:04d}"

        # Convert donation_date (date) to datetime for entry_date
        if isinstance(donation.donation_date, date):
            entry_date = datetime.combine(donation.donation_date, datetime.min.time())
        else:
            entry_date = donation.donation_date

        # Create journal entry
        # Note: created_by is required, so use 1 as default if None (system user)
        created_by = donation.created_by if donation.created_by else 1
        
        journal_entry = JournalEntry(
            temple_id=temple_id,
            entry_date=entry_date,
            entry_number=entry_number,
            narration=narration,
            reference_type=TransactionType.DONATION,
            reference_id=donation.id,
            total_amount=donation.amount,
            status=JournalEntryStatus.POSTED,
            created_by=created_by,
            posted_by=created_by,
            posted_at=datetime.utcnow()
        )
        db.add(journal_entry)
        db.flush()  # Get journal_entry.id

        # Create journal lines
        # Debit: Payment Account (Cash/Bank) or Asset Account (In-kind)
        if donation.donation_type == DonationType.IN_KIND:
            debit_description = f"In-kind donation: {donation.item_name or 'Item'}"
            if donation.quantity and donation.unit:
                debit_description += f" - {donation.quantity} {donation.unit}"
        else:
            debit_description = f"Donation received via {donation.payment_mode or 'Cash'}"
        
        debit_line = JournalLine(
            journal_entry_id=journal_entry.id,
            account_id=debit_account.id,
            debit_amount=donation.amount,
            credit_amount=0,
            description=debit_description
        )

        # Credit: Donation Income (Income increases)
        credit_description = f"Donation income - {donation.category.name if donation.category else 'General'}"
        if donation.donation_type == DonationType.IN_KIND:
            credit_description += f" (In-kind: {donation.item_name or 'Item'})"
        
        credit_line = JournalLine(
            journal_entry_id=journal_entry.id,
            account_id=credit_account.id,
            debit_amount=0,
            credit_amount=donation.amount,
            description=credit_description
        )

        db.add(debit_line)
        db.add(credit_line)

        return journal_entry

    except Exception as e:
        print(f"Error posting donation to accounting: {str(e)}")
        return None


# Pydantic Schemas
class DonationBase(BaseModel):
    devotee_name: str
    devotee_phone: str
    amount: float
    category: str
    donation_type: DonationType = DonationType.CASH  # Default to cash donation
    payment_mode: Optional[str] = "Cash"  # Required for cash donations, optional for in-kind
    address: Optional[str] = None  # Street address
    pincode: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    country: Optional[str] = "India"
    
    # In-Kind Donation Fields (only for donation_type = IN_KIND)
    inkind_subtype: Optional[InKindDonationSubType] = None
    item_name: Optional[str] = None
    item_description: Optional[str] = None
    quantity: Optional[float] = None
    unit: Optional[str] = None
    value_assessed: Optional[float] = None  # Assessed value (same as amount for in-kind)
    appraised_by: Optional[str] = None
    appraisal_date: Optional[date] = None
    
    # For Precious Items
    purity: Optional[str] = None
    weight_gross: Optional[float] = None
    weight_net: Optional[float] = None
    
    # For Event Sponsorship
    event_name: Optional[str] = None
    event_date: Optional[date] = None
    sponsorship_category: Optional[str] = None
    
    # Links to Inventory or Asset
    inventory_item_id: Optional[int] = None
    asset_id: Optional[int] = None
    store_id: Optional[int] = None  # Store where inventory is received
    
    # Photo/Document URLs
    photo_url: Optional[str] = None
    document_url: Optional[str] = None


class DonationCreate(DonationBase):
    notes: Optional[str] = None


class DonationResponse(BaseModel):
    id: int
    receipt_number: str
    amount: float
    donation_type: DonationType
    payment_mode: Optional[str] = None
    donation_date: date
    devotee: Optional[dict] = None
    category: Optional[dict] = None
    
    # In-Kind Donation Fields
    inkind_subtype: Optional[InKindDonationSubType] = None
    item_name: Optional[str] = None
    item_description: Optional[str] = None
    quantity: Optional[float] = None
    unit: Optional[str] = None
    value_assessed: Optional[float] = None
    
    created_at: str
    
    class Config:
        from_attributes = True


@router.post("/", response_model=dict, status_code=status.HTTP_201_CREATED)
def create_donation(
    donation: DonationCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    request: Request = None
):
    """Create a new donation"""
    # Find or create devotee
    devotee = db.query(Devotee).filter(Devotee.phone == donation.devotee_phone).first()
    if not devotee:
        # Create devotee if doesn't exist
        devotee = Devotee(
            name=donation.devotee_name,
            full_name=donation.devotee_name,
            phone=donation.devotee_phone,
            address=donation.address,
            pincode=donation.pincode,
            city=donation.city,
            state=donation.state,
            country=donation.country or "India",
            temple_id=current_user.temple_id if current_user else None
        )
        db.add(devotee)
        db.flush()
    else:
        # Update devotee info if provided
        if donation.address and not devotee.address:
            devotee.address = donation.address
        if donation.pincode and not devotee.pincode:
            devotee.pincode = donation.pincode
        if donation.city and not devotee.city:
            devotee.city = donation.city
        if donation.state and not devotee.state:
            devotee.state = donation.state
        if donation.country and not devotee.country:
            devotee.country = donation.country
        if donation.devotee_name and devotee.name != donation.devotee_name:
            devotee.name = donation.devotee_name
            devotee.full_name = donation.devotee_name
        db.flush()
    
    # Find or create category
    category = db.query(DonationCategory).filter(
        DonationCategory.name == donation.category
    ).first()
    if not category:
        category = DonationCategory(
            name=donation.category,
            temple_id=current_user.temple_id if current_user else None
        )
        db.add(category)
        db.flush()
    
    # Generate receipt number
    year = datetime.now().year
    last_donation = db.query(Donation).filter(
        func.extract('year', Donation.donation_date) == year
    ).order_by(Donation.id.desc()).first()
    
    seq = 1
    if last_donation and last_donation.receipt_number:
        try:
            seq = int(last_donation.receipt_number.split('-')[-1]) + 1
        except:
            seq = 1
    
    receipt_number = f"TMP001-{year}-{str(seq).zfill(5)}"
    
    # Validate in-kind donation fields if donation_type is IN_KIND
    if donation.donation_type == DonationType.IN_KIND:
        if not donation.item_name:
            raise HTTPException(status_code=400, detail="item_name is required for in-kind donations")
        if not donation.quantity or donation.quantity <= 0:
            raise HTTPException(status_code=400, detail="quantity must be greater than 0 for in-kind donations")
        if not donation.unit:
            raise HTTPException(status_code=400, detail="unit is required for in-kind donations")
        if not donation.inkind_subtype:
            raise HTTPException(status_code=400, detail="inkind_subtype is required for in-kind donations")
        # Use value_assessed if provided, otherwise use amount
        assessed_value = donation.value_assessed if donation.value_assessed is not None else donation.amount
    else:
        # For cash donations, payment_mode is required
        if not donation.payment_mode:
            raise HTTPException(status_code=400, detail="payment_mode is required for cash donations")
        assessed_value = None
    
    # Duplicate detection: Check for similar donation within last 5 minutes
    duplicate_window = datetime.now() - timedelta(minutes=5)
    duplicate_check = db.query(Donation).filter(
        Donation.devotee_id == devotee.id,
        Donation.amount == donation.amount,
        Donation.donation_date >= duplicate_window.date() if isinstance(duplicate_window, datetime) else date.today(),
        Donation.is_cancelled == False
    ).first()
    
    if duplicate_check:
        # Check if within 5 minutes
        if duplicate_check.donation_date == date.today():
            raise HTTPException(
                status_code=400,
                detail=f"Possible duplicate donation detected. Similar donation (₹{donation.amount}) from {devotee.name} was recorded today. Receipt: {duplicate_check.receipt_number}. Please verify before proceeding."
            )
    
    # Create donation
    db_donation = Donation(
        temple_id=current_user.temple_id if current_user else None,
        devotee_id=devotee.id,
        category_id=category.id,
        receipt_number=receipt_number,
        donation_type=donation.donation_type,
        amount=donation.amount,
        payment_mode=donation.payment_mode if donation.donation_type == DonationType.CASH else None,
        donation_date=date.today(),
        financial_year=f"{year}-{str(year+1)[-2:]}",
        notes=donation.notes,
        created_by=current_user.id if current_user else None,
        # In-kind donation fields
        inkind_subtype=donation.inkind_subtype if donation.donation_type == DonationType.IN_KIND else None,
        item_name=donation.item_name if donation.donation_type == DonationType.IN_KIND else None,
        item_description=donation.item_description if donation.donation_type == DonationType.IN_KIND else None,
        quantity=donation.quantity if donation.donation_type == DonationType.IN_KIND else None,
        unit=donation.unit if donation.donation_type == DonationType.IN_KIND else None,
        value_assessed=assessed_value if donation.donation_type == DonationType.IN_KIND else None,
        appraised_by=donation.appraised_by if donation.donation_type == DonationType.IN_KIND else None,
        appraisal_date=donation.appraisal_date if donation.donation_type == DonationType.IN_KIND else None,
        purity=donation.purity if donation.donation_type == DonationType.IN_KIND else None,
        weight_gross=donation.weight_gross if donation.donation_type == DonationType.IN_KIND else None,
        weight_net=donation.weight_net if donation.donation_type == DonationType.IN_KIND else None,
        event_name=donation.event_name if donation.donation_type == DonationType.IN_KIND else None,
        event_date=donation.event_date if donation.donation_type == DonationType.IN_KIND else None,
        sponsorship_category=donation.sponsorship_category if donation.donation_type == DonationType.IN_KIND else None,
        inventory_item_id=donation.inventory_item_id if donation.donation_type == DonationType.IN_KIND else None,
        asset_id=donation.asset_id if donation.donation_type == DonationType.IN_KIND else None,
        store_id=donation.store_id if donation.donation_type == DonationType.IN_KIND else None,
        photo_url=donation.photo_url if donation.donation_type == DonationType.IN_KIND else None,
        document_url=donation.document_url if donation.donation_type == DonationType.IN_KIND else None,
        current_balance=donation.quantity if (donation.donation_type == DonationType.IN_KIND and donation.inkind_subtype == InKindDonationSubType.INVENTORY) else None
    )
    db.add(db_donation)
    db.flush()  # Flush to get the ID
    
    # If in-kind donation is inventory type, create stock movement
    if donation.donation_type == DonationType.IN_KIND and donation.inkind_subtype == InKindDonationSubType.INVENTORY:
        from app.models.inventory import StockMovement, StockMovementType, StockBalance
        from datetime import datetime as dt
        
        # Get or create inventory item if inventory_item_id is provided
        if donation.inventory_item_id:
            item_id = donation.inventory_item_id
        else:
            # Create a new inventory item for this donation
            from app.models.inventory import Item, ItemCategory, Unit as InvUnit
            # Try to find existing item by name
            existing_item = db.query(Item).filter(
                Item.temple_id == current_user.temple_id if current_user else None,
                Item.name.ilike(donation.item_name)
            ).first()
            
            if existing_item:
                item_id = existing_item.id
            else:
                # Create new item
                new_item = Item(
                    temple_id=current_user.temple_id if current_user else None,
                    code=f"ITM-{db_donation.id:05d}",
                    name=donation.item_name,
                    description=donation.item_description or f"In-kind donation: {donation.item_name}",
                    category=ItemCategory.GROCERY,  # Default to grocery for consumables
                    unit=InvUnit.KG if donation.unit.lower() in ['kg', 'kilogram'] else InvUnit.PIECE
                )
                db.add(new_item)
                db.flush()
                item_id = new_item.id
        
        # Get default store if store_id not provided
        store_id = donation.store_id
        if not store_id:
            from app.models.inventory import Store
            default_store = db.query(Store).filter(
                Store.temple_id == current_user.temple_id if current_user else None,
                Store.is_active == True
            ).first()
            if default_store:
                store_id = default_store.id
            else:
                # Create default store if none exists
                default_store = Store(
                    temple_id=current_user.temple_id if current_user else None,
                    code="ST001",
                    name="Main Store",
                    is_active=True
                )
                db.add(default_store)
                db.flush()
                store_id = default_store.id
        
        # Create stock movement for donation receipt
        movement_number = f"DON/{year}/{db_donation.id:05d}"
        stock_movement = StockMovement(
            temple_id=current_user.temple_id if current_user else None,
            movement_type=StockMovementType.PURCHASE,  # Treat donation as purchase
            movement_number=movement_number,
            movement_date=date.today(),
            item_id=item_id,
            store_id=store_id,
            quantity=donation.quantity,
            unit_price=assessed_value / donation.quantity if donation.quantity > 0 else 0,
            total_value=assessed_value,
            reference_number=db_donation.receipt_number,
            notes=f"In-kind donation from {devotee.name}",
            created_by=current_user.id if current_user else None
        )
        db.add(stock_movement)
        db.flush()
        
        # Update or create stock balance
        stock_balance = db.query(StockBalance).filter(
            StockBalance.item_id == item_id,
            StockBalance.store_id == store_id
        ).first()
        
        if stock_balance:
            stock_balance.quantity += donation.quantity
            stock_balance.value += assessed_value
            stock_balance.last_movement_date = date.today()
            stock_balance.last_movement_id = stock_movement.id
        else:
            stock_balance = StockBalance(
                temple_id=current_user.temple_id if current_user else None,
                item_id=item_id,
                store_id=store_id,
                quantity=donation.quantity,
                value=assessed_value,
                last_movement_date=date.today(),
                last_movement_id=stock_movement.id
            )
            db.add(stock_balance)
        
        # Update donation with inventory_item_id and store_id
        db_donation.inventory_item_id = item_id
        db_donation.store_id = store_id
    
    db.commit()
    db.refresh(db_donation)

    # Post to accounting system
    journal_entry = post_donation_to_accounting(db, db_donation, current_user.temple_id if current_user else None)
    if journal_entry:
        db.commit()  # Commit the journal entry
        message = "Donation recorded successfully and posted to accounting"
    else:
        message = "Donation recorded but accounting entry failed. Please check chart of accounts."

    # Audit log
    log_action(
        db=db,
        user=current_user,
        action="CREATE_DONATION",
        entity_type="Donation",
        entity_id=db_donation.id,
        new_values=get_entity_dict(db_donation),
        description=f"Created donation: ₹{db_donation.amount} from {devotee.name} (Receipt: {db_donation.receipt_number})",
        ip_address=request.client.host if request else None,
        user_agent=request.headers.get("user-agent") if request else None
    )

    # Auto-send SMS/Email receipt (if enabled and devotee preferences allow)
    try:
        if devotee.receive_sms and devotee.phone:
            # Send SMS receipt (async - don't block response)
            # TODO: Implement SMS service integration
            pass
        if devotee.receive_email and devotee.email:
            # Send Email receipt (async - don't block response)
            # TODO: Implement Email service integration
            pass
    except Exception as e:
        # Don't fail donation creation if SMS/Email fails
        print(f"Failed to send receipt notification: {str(e)}")

    return {
        "id": db_donation.id,
        "receipt_number": db_donation.receipt_number,
        "amount": db_donation.amount,
        "journal_entry": journal_entry.entry_number if journal_entry else "NOT_POSTED",
        "message": message,
        "accounting_posted": journal_entry is not None
    }


@router.get("/{donation_id}/receipt/pdf")
def get_donation_receipt_pdf(
    donation_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Generate PDF receipt for a single donation
    Professional receipt format with temple details and 80G information
    """
    # Get donation
    donation = db.query(Donation).filter(Donation.id == donation_id).first()
    if not donation:
        raise HTTPException(status_code=404, detail="Donation not found")
    
    # Get temple info
    temple = None
    temple_logo_path = None
    if donation.temple_id:
        temple = db.query(Temple).filter(Temple.id == donation.temple_id).first()
        if temple and temple.logo_url:
            try:
                if temple.logo_url.startswith('http'):
                    response = requests.get(temple.logo_url, timeout=5)
                    if response.status_code == 200:
                        temple_logo_path = io.BytesIO(response.content)
                elif os.path.exists(temple.logo_url):
                    temple_logo_path = temple.logo_url
            except:
                temple_logo_path = None
    
    # Get devotee info
    devotee = donation.devotee
    category = donation.category
    
    # Create PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    
    styles = getSampleStyleSheet()
    
    # Title style
    title_style = ParagraphStyle(
        'ReceiptTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#FF9933'),
        spaceAfter=12,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    # Header style
    header_style = ParagraphStyle(
        'ReceiptHeader',
        parent=styles['Normal'],
        fontSize=14,
        textColor=colors.black,
        alignment=TA_CENTER,
        spaceAfter=6
    )
    
    # Temple header
    if temple:
        if temple_logo_path:
            try:
                logo = Image(temple_logo_path, width=1.2*inch, height=1.2*inch)
                logo.hAlign = 'CENTER'
                elements.append(logo)
                elements.append(Spacer(1, 0.1*inch))
            except:
                pass
        
        if temple.name:
            elements.append(Paragraph(temple.name, title_style))
        
        if temple.address:
            elements.append(Paragraph(temple.address, header_style))
        
        if temple.phone:
            elements.append(Paragraph(f"Phone: {temple.phone}", styles['Normal']))
        
        if temple.email:
            elements.append(Paragraph(f"Email: {temple.email}", styles['Normal']))
        
        elements.append(Spacer(1, 0.2*inch))
        elements.append(Paragraph("_" * 80, styles['Normal']))
        elements.append(Spacer(1, 0.2*inch))
    
    # Receipt title
    elements.append(Paragraph("DONATION RECEIPT", title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Receipt details table
    receipt_data = [
        ["Receipt Number:", donation.receipt_number],
        ["Date:", donation.donation_date.strftime('%d-%m-%Y') if donation.donation_date else ""],
        ["Devotee Name:", devotee.name if devotee else "Anonymous"],
        ["Phone:", devotee.phone if devotee else "N/A"],
        ["Address:", devotee.address if devotee and devotee.address else "N/A"],
        ["Category:", category.name if category else "N/A"],
        ["Payment Mode:", donation.payment_mode.upper()],
        ["Amount:", f"₹ {donation.amount:,.2f}"]
    ]
    
    # Add 80G information if applicable
    if category and category.is_80g_eligible and temple and temple.certificate_80g_number:
        receipt_data.append(["80G Certificate:", f"Yes - {temple.certificate_80g_number}"])
        receipt_data.append(["80G Valid From:", temple.certificate_80g_valid_from or "N/A"])
        receipt_data.append(["80G Valid To:", temple.certificate_80g_valid_to or "N/A"])
    
    receipt_table = Table(receipt_data, colWidths=[2.5*inch, 4*inch])
    receipt_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    
    elements.append(receipt_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Amount in words (simple conversion)
    amount_words = f"Rupees {_number_to_words(int(donation.amount))} Only"
    elements.append(Paragraph(f"<b>Amount in Words:</b> {amount_words}", styles['Normal']))
    elements.append(Spacer(1, 0.3*inch))
    
    # Footer
    footer_style = ParagraphStyle(
        'ReceiptFooter',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.grey,
        alignment=TA_CENTER
    )
    
    elements.append(Spacer(1, 0.5*inch))
    elements.append(Paragraph("_" * 80, styles['Normal']))
    elements.append(Spacer(1, 0.1*inch))
    
    if temple and temple.authorized_signatory_name:
        elements.append(Paragraph(f"Authorized Signatory: {temple.authorized_signatory_name}", styles['Normal']))
        if temple.authorized_signatory_designation:
            elements.append(Paragraph(temple.authorized_signatory_designation, styles['Normal']))
    
    elements.append(Spacer(1, 0.2*inch))
    elements.append(Paragraph(f"Generated on {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}", footer_style))
    elements.append(Paragraph("MandirSync Temple Management System", footer_style))
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"receipt_{donation.receipt_number}.pdf"
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"inline; filename={filename}"}
    )


def _number_to_words(n):
    """Convert number to words (simple implementation)"""
    ones = ["", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine"]
    tens = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]
    teens = ["Ten", "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen", "Sixteen", "Seventeen", "Eighteen", "Nineteen"]
    
    if n == 0:
        return "Zero"
    
    def convert_hundreds(num):
        result = ""
        if num >= 100:
            result += ones[num // 100] + " Hundred "
            num %= 100
        if num >= 20:
            result += tens[num // 10] + " "
            num %= 10
        elif num >= 10:
            result += teens[num - 10] + " "
            return result
        if num > 0:
            result += ones[num] + " "
        return result
    
    result = ""
    if n >= 10000000:  # Crores
        result += convert_hundreds(n // 10000000) + "Crore "
        n %= 10000000
    if n >= 100000:  # Lakhs
        result += convert_hundreds(n // 100000) + "Lakh "
        n %= 100000
    if n >= 1000:  # Thousands
        result += convert_hundreds(n // 1000) + "Thousand "
        n %= 1000
    if n > 0:
        result += convert_hundreds(n)
    
    return result.strip()


@router.get("/", response_model=List[DonationResponse])
def get_donations(
    skip: int = 0,
    limit: int = 100,
    date: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get list of donations"""
    query = db.query(Donation)
    
    if date:
        query = query.filter(Donation.donation_date == date)
    elif date_from and date_to:
        query = query.filter(
            and_(
                Donation.donation_date >= date_from,
                Donation.donation_date <= date_to
            )
        )
    
    donations = query.order_by(Donation.donation_date.desc()).offset(skip).limit(limit).all()
    
    # Format response
    result = []
    for d in donations:
        result.append({
            "id": d.id,
            "receipt_number": d.receipt_number,
            "amount": d.amount,
            "payment_mode": d.payment_mode,
            "donation_date": d.donation_date,
            "devotee": {
                "id": d.devotee.id if d.devotee else None,
                "name": d.devotee.name if d.devotee else None,
                "phone": d.devotee.phone if d.devotee else None,
                "email": d.devotee.email if d.devotee else None,
                "address": d.devotee.address if d.devotee else None
            } if d.devotee else None,
            "devotee_phone": d.devotee.phone if d.devotee else None,  # For backward compatibility
            "category": {
                "id": d.category.id if d.category else None,
                "name": d.category.name if d.category else None
            } if d.category else None,
            "created_at": d.created_at
        })
    
    return result


@router.get("/report/daily")
def get_daily_report(
    date: Optional[str] = Query(None, description="Date in YYYY-MM-DD format"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get daily donation report"""
    from datetime import date as date_class
    report_date = date if date else date_class.today().isoformat()
    
    donations = db.query(Donation).filter(
        Donation.donation_date == report_date
    ).all()
    
    total = sum(d.amount for d in donations)
    
    # Group by category
    by_category = {}
    for d in donations:
        cat_name = d.category.name if d.category else "Unknown"
        if cat_name not in by_category:
            by_category[cat_name] = {"amount": 0, "count": 0}
        by_category[cat_name]["amount"] += d.amount
        by_category[cat_name]["count"] += 1
    
    return {
        "date": report_date,
        "total": total,
        "count": len(donations),
        "by_category": [
            {"category": k, "amount": v["amount"], "count": v["count"]}
            for k, v in by_category.items()
        ]
    }


@router.get("/report/monthly")
def get_monthly_report(
    month: int = Query(..., ge=1, le=12),
    year: int = Query(..., ge=2020),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get monthly donation report"""
    donations = db.query(Donation).filter(
        Donation.temple_id == current_user.temple_id,
        func.extract('year', Donation.donation_date) == year,
        func.extract('month', Donation.donation_date) == month,
        Donation.is_cancelled == False
    ).all()
    
    total = sum(d.amount for d in donations)
    
    # Group by category
    by_category = {}
    for d in donations:
        cat_name = d.category.name if d.category else "Unknown"
        if cat_name not in by_category:
            by_category[cat_name] = {"amount": 0, "count": 0}
        by_category[cat_name]["amount"] += d.amount
        by_category[cat_name]["count"] += 1
    
    return {
        "month": month,
        "year": year,
        "total": total,
        "count": len(donations),
        "by_category": [
            {"category": k, "amount": v["amount"], "count": v["count"]}
            for k, v in by_category.items()
        ]
    }


@router.get("/report/category-wise")
def get_category_wise_report(
    date_from: Optional[str] = Query(None, description="Start date (YYYY-MM-DD). Default: today"),
    date_to: Optional[str] = Query(None, description="End date (YYYY-MM-DD). Default: today"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Get category-wise donation report for a date range
    Default: Today's donations grouped by category
    """
    from datetime import date as date_class
    
    # Default to today if not specified
    if not date_from:
        date_from = date_class.today().isoformat()
    if not date_to:
        date_to = date_from
    
    # Parse dates
    start_date = datetime.strptime(date_from, '%Y-%m-%d').date()
    end_date = datetime.strptime(date_to, '%Y-%m-%d').date()
    
    donations = db.query(Donation).filter(
        Donation.temple_id == current_user.temple_id,
        Donation.donation_date >= start_date,
        Donation.donation_date <= end_date,
        Donation.is_cancelled == False
    ).all()
    
    total = sum(d.amount for d in donations)
    
    # Group by category
    by_category = {}
    for d in donations:
        cat_name = d.category.name if d.category else "Unknown"
        if cat_name not in by_category:
            by_category[cat_name] = {"amount": 0, "count": 0}
        by_category[cat_name]["amount"] += d.amount
        by_category[cat_name]["count"] += 1
    
    return {
        "date_from": date_from,
        "date_to": date_to,
        "total": total,
        "count": len(donations),
        "by_category": [
            {"category": k, "amount": v["amount"], "count": v["count"]}
            for k, v in sorted(by_category.items(), key=lambda x: x[1]["amount"], reverse=True)
        ]
    }


@router.get("/report/detailed")
def get_detailed_donation_report(
    date_from: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    date_to: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    category: Optional[str] = Query(None, description="Filter by category"),
    payment_mode: Optional[str] = Query(None, description="Filter by payment mode"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Get detailed donation report with:
    - Date
    - Devotee Name
    - Mobile Number
    - Donation Category
    - Payment Mode
    - Amount
    """
    from datetime import date as date_class
    
    query = db.query(Donation).filter(
        Donation.temple_id == current_user.temple_id,
        Donation.is_cancelled == False
    )
    
    # Apply date filters
    if date_from:
        start_date = datetime.strptime(date_from, '%Y-%m-%d').date()
        query = query.filter(Donation.donation_date >= start_date)
    
    if date_to:
        end_date = datetime.strptime(date_to, '%Y-%m-%d').date()
        query = query.filter(Donation.donation_date <= end_date)
    
    # Apply category filter
    if category:
        query = query.join(DonationCategory).filter(DonationCategory.name == category)
    
    # Apply payment mode filter
    if payment_mode:
        query = query.filter(Donation.payment_mode.ilike(f'%{payment_mode}%'))
    
    donations = query.order_by(Donation.donation_date.desc(), Donation.id.desc()).all()
    
    result = []
    for d in donations:
        result.append({
            "id": d.id,
            "receipt_number": d.receipt_number,
            "date": d.donation_date.isoformat() if d.donation_date else None,
            "devotee_name": d.devotee.name if d.devotee else "Anonymous",
            "mobile_number": d.devotee.phone if d.devotee else None,
            "category": d.category.name if d.category else "Unknown",
            "payment_mode": d.payment_mode,
            "amount": float(d.amount),
            "transaction_id": d.transaction_id,
            "notes": d.notes
        })
    
    return {
        "date_from": date_from,
        "date_to": date_to,
        "filters": {
            "category": category,
            "payment_mode": payment_mode
        },
        "total": sum(d.amount for d in donations),
        "count": len(donations),
        "donations": result
    }


@router.get("/export/pdf")
def export_donations_pdf(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Export donations to PDF format"""
    # Get temple info
    temple = None
    temple_logo_path = None
    if current_user and current_user.temple_id:
        temple = db.query(Temple).filter(Temple.id == current_user.temple_id).first()
        if temple and temple.logo_url:
            # Try to download logo if it's a URL, or use local path
            try:
                if temple.logo_url.startswith('http'):
                    # Download logo
                    response = requests.get(temple.logo_url, timeout=5)
                    if response.status_code == 200:
                        logo_data = io.BytesIO(response.content)
                        temple_logo_path = logo_data
                else:
                    # Local file path
                    if os.path.exists(temple.logo_url):
                        temple_logo_path = temple.logo_url
            except:
                temple_logo_path = None
    
    query = db.query(Donation)
    
    if date_from and date_to:
        query = query.filter(
            and_(
                Donation.donation_date >= date_from,
                Donation.donation_date <= date_to
            )
        )
    
    donations = query.order_by(Donation.donation_date.desc()).all()
    
    # Create PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#FF9933'),
        spaceAfter=12,
        alignment=TA_CENTER
    )
    
    # Temple header
    if temple:
        if temple_logo_path:
            try:
                logo = Image(temple_logo_path, width=1*inch, height=1*inch)
                logo.hAlign = 'CENTER'
                elements.append(logo)
                elements.append(Spacer(1, 0.1*inch))
            except:
                pass
        
        if temple.name:
            elements.append(Paragraph(temple.name, title_style))
        
        if temple.address:
            elements.append(Paragraph(temple.address, styles['Normal']))
            elements.append(Spacer(1, 0.1*inch))
        
        if temple.phone:
            elements.append(Paragraph(f"Phone: {temple.phone}", styles['Normal']))
            elements.append(Spacer(1, 0.2*inch))
    
    # Report title
    report_title = "DONATION REPORT"
    if date_from and date_to:
        report_title += f"<br/><font size=10>Period: {date_from} to {date_to}</font>"
    elements.append(Paragraph(report_title, title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Table data
    table_data = [["Receipt", "Date", "Devotee", "Phone", "Amount (₹)", "Category", "Payment"]]
    
    total_amount = 0
    for d in donations:
        table_data.append([
            d.receipt_number or "",
            d.donation_date.strftime('%d-%m-%Y') if d.donation_date else "",
            d.devotee.name if d.devotee else "",
            d.devotee.phone if d.devotee else "",
            f"{d.amount:,.0f}",
            d.category.name if d.category else "",
            d.payment_mode or ""
        ])
        total_amount += d.amount
    
    # Total row
    table_data.append([
        "TOTAL", "", "", "", f"<b>{total_amount:,.0f}</b>", f"{len(donations)} donations", ""
    ])
    
    # Create table
    table = Table(table_data, colWidths=[1.2*inch, 0.9*inch, 1.2*inch, 1*inch, 0.9*inch, 1.2*inch, 0.8*inch])
    table.setStyle(TableStyle([
        # Header row
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FF9933')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        # Data rows
        ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -2), colors.black),
        ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -2), 9),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        # Total row
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 10),
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Footer
    footer_text = f"Generated on {datetime.now().strftime('%d-%m-%Y %H:%M:%S')} | MandirSync Temple Management System"
    elements.append(Paragraph(footer_text, styles['Normal']))
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"donations_{date_from or 'all'}_{date_to or date.today()}.pdf"
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/excel")
def export_donations_excel(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Export donations to Excel format"""
    # Get temple info
    temple = None
    temple_logo_path = None
    if current_user and current_user.temple_id:
        temple = db.query(Temple).filter(Temple.id == current_user.temple_id).first()
        if temple and temple.logo_url:
            # Try to download logo if it's a URL, or use local path
            try:
                if temple.logo_url.startswith('http'):
                    # Download logo
                    response = requests.get(temple.logo_url, timeout=5)
                    if response.status_code == 200:
                        logo_data = io.BytesIO(response.content)
                        temple_logo_path = logo_data
                else:
                    # Local file path
                    if os.path.exists(temple.logo_url):
                        temple_logo_path = temple.logo_url
            except:
                temple_logo_path = None
    
    query = db.query(Donation)
    
    if date_from and date_to:
        query = query.filter(
            and_(
                Donation.donation_date >= date_from,
                Donation.donation_date <= date_to
            )
        )
    
    donations = query.order_by(Donation.donation_date.desc()).all()
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Donations Report"
    
    # Styles
    header_fill = PatternFill(start_color="FF9933", end_color="FF9933", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=16, color="FF9933")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    row = 1
    
    # Temple header with logo
    if temple:
        # Add logo if available
        if temple_logo_path:
            try:
                logo = ExcelImage(temple_logo_path)
                logo.width = 100
                logo.height = 100
                ws.add_image(logo, 'A1')
                row = 4  # Start after logo
            except Exception as e:
                # Logo failed, continue without it
                row = 1
        
        ws.merge_cells(f'A{row}:G{row}')
        cell = ws.cell(row=row, column=1)
        cell.value = temple.name or "Temple Donations Report"
        cell.font = title_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 1
        
        if temple.address:
            ws.merge_cells(f'A{row}:G{row}')
            cell = ws.cell(row=row, column=1)
            cell.value = temple.address
            cell.alignment = Alignment(horizontal='center', vertical='center')
            row += 1
        
        if temple.phone:
            ws.merge_cells(f'A{row}:G{row}')
            cell = ws.cell(row=row, column=1)
            cell.value = f"Phone: {temple.phone}"
            cell.alignment = Alignment(horizontal='center', vertical='center')
            row += 1
    
    # Report period
    if date_from and date_to:
        ws.merge_cells(f'A{row}:G{row}')
        cell = ws.cell(row=row, column=1)
        cell.value = f"Period: {date_from} to {date_to}"
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True)
        row += 1
    
    row += 1  # Empty row
    
    # Headers
    headers = ["Receipt Number", "Date", "Devotee Name", "Phone", "Amount (₹)", "Category", "Payment Mode"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    row += 1
    
    # Data rows
    total_amount = 0
    for d in donations:
        ws.cell(row=row, column=1, value=d.receipt_number).border = border
        ws.cell(row=row, column=2, value=d.donation_date.strftime('%Y-%m-%d') if d.donation_date else "").border = border
        ws.cell(row=row, column=3, value=d.devotee.name if d.devotee else "").border = border
        ws.cell(row=row, column=4, value=d.devotee.phone if d.devotee else "").border = border
        ws.cell(row=row, column=5, value=d.amount).border = border
        ws.cell(row=row, column=6, value=d.category.name if d.category else "").border = border
        ws.cell(row=row, column=7, value=d.payment_mode).border = border
        total_amount += d.amount
        row += 1
    
    # Total row
    row += 1
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws.cell(row=row, column=1)
    cell.value = "TOTAL"
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal='right', vertical='center')
    cell.border = border
    
    ws.cell(row=row, column=5, value=total_amount).font = Font(bold=True, size=12)
    ws.cell(row=row, column=5).border = border
    ws.cell(row=row, column=6, value=f"{len(donations)} donations").border = border
    ws.cell(row=row, column=7).border = border
    
    # Auto-adjust column widths
    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"donations_{date_from or 'all'}_{date_to or date.today()}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.post("/bulk-import", response_model=dict)
async def bulk_import_donations(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    request: Request = None
):
    """
    Bulk import donations from CSV/Excel file
    Expected format:
    - CSV: devotee_name, devotee_phone, amount, category, payment_mode, address, city, state, pincode, notes
    - Excel: Same columns
    """
    temple_id = current_user.temple_id if current_user else None
    
    # Read file
    contents = await file.read()
    
    # Determine file type
    file_extension = file.filename.split('.')[-1].lower() if file.filename else 'csv'
    
    donations_data = []
    errors = []
    success_count = 0
    
    try:
        if file_extension in ['xlsx', 'xls']:
            # Read Excel
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(contents))
            ws = wb.active
            
            # Read header row
            headers = [cell.value for cell in ws[1]]
            
            # Read data rows
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):  # Skip empty rows
                    continue
                
                row_dict = dict(zip(headers, row))
                donations_data.append({
                    'row': row_idx,
                    'data': row_dict
                })
        else:
            # Read CSV
            csv_data = io.StringIO(contents.decode('utf-8'))
            reader = csv.DictReader(csv_data)
            
            for row_idx, row in enumerate(reader, start=2):
                donations_data.append({
                    'row': row_idx,
                    'data': row
                })
        
        # Process each donation
        for item in donations_data:
            row_num = item['row']
            row_data = item['data']
            
            try:
                # Validate required fields
                if not row_data.get('devotee_name') or not row_data.get('devotee_phone') or not row_data.get('amount'):
                    errors.append(f"Row {row_num}: Missing required fields (devotee_name, devotee_phone, amount)")
                    continue
                
                # Create donation using existing create_donation logic
                donation_create = DonationBase(
                    devotee_name=str(row_data.get('devotee_name', '')).strip(),
                    devotee_phone=str(row_data.get('devotee_phone', '')).strip(),
                    amount=float(row_data.get('amount', 0)),
                    category=str(row_data.get('category', 'General Donation')).strip(),
                    payment_mode=str(row_data.get('payment_mode', 'Cash')).strip(),
                    address=str(row_data.get('address', '')).strip() if row_data.get('address') else None,
                    city=str(row_data.get('city', '')).strip() if row_data.get('city') else None,
                    state=str(row_data.get('state', '')).strip() if row_data.get('state') else None,
                    pincode=str(row_data.get('pincode', '')).strip() if row_data.get('pincode') else None,
                    country=str(row_data.get('country', 'India')).strip() if row_data.get('country') else 'India',
                    notes=str(row_data.get('notes', '')).strip() if row_data.get('notes') else None
                )
                
                # Call create_donation endpoint logic (inline)
                # Find or create devotee
                devotee = db.query(Devotee).filter(Devotee.phone == donation_create.devotee_phone).first()
                if not devotee:
                    devotee = Devotee(
                        name=donation_create.devotee_name,
                        full_name=donation_create.devotee_name,
                        phone=donation_create.devotee_phone,
                        address=donation_create.address,
                        pincode=donation_create.pincode,
                        city=donation_create.city,
                        state=donation_create.state,
                        country=donation_create.country or "India",
                        temple_id=temple_id
                    )
                    db.add(devotee)
                    db.flush()
                
                # Find or create category
                category = db.query(DonationCategory).filter(
                    DonationCategory.name == donation_create.category
                ).first()
                if not category:
                    category = DonationCategory(
                        name=donation_create.category,
                        is_80g_eligible=True,
                        temple_id=temple_id
                    )
                    db.add(category)
                    db.flush()
                
                # Generate receipt number
                year = datetime.now().year
                last_donation = db.query(Donation).filter(
                    func.extract('year', Donation.donation_date) == year
                ).order_by(Donation.id.desc()).first()
                
                seq = 1
                if last_donation and last_donation.receipt_number:
                    try:
                        seq = int(last_donation.receipt_number.split('-')[-1]) + 1
                    except:
                        seq = 1
                
                receipt_number = f"TMP001-{year}-{str(seq).zfill(5)}"
                
                # Duplicate check
                duplicate_check = db.query(Donation).filter(
                    Donation.devotee_id == devotee.id,
                    Donation.amount == donation_create.amount,
                    Donation.donation_date == date.today(),
                    Donation.is_cancelled == False
                ).first()
                
                if duplicate_check:
                    errors.append(f"Row {row_num}: Possible duplicate - Similar donation exists (Receipt: {duplicate_check.receipt_number})")
                    continue
                
                # Create donation
                db_donation = Donation(
                    temple_id=temple_id,
                    devotee_id=devotee.id,
                    category_id=category.id,
                    receipt_number=receipt_number,
                    donation_type=DonationType.CASH,
                    amount=donation_create.amount,
                    payment_mode=donation_create.payment_mode,
                    donation_date=date.today(),
                    financial_year=f"{year}-{str(year+1)[-2:]}",
                    notes=donation_create.notes,
                    created_by=current_user.id if current_user else None
                )
                db.add(db_donation)
                db.flush()
                
                # Post to accounting
                journal_entry = post_donation_to_accounting(db, db_donation, temple_id)
                if journal_entry:
                    db.commit()
                
                success_count += 1
                
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
                db.rollback()
        
        return {
            "success": True,
            "total_rows": len(donations_data),
            "success_count": success_count,
            "error_count": len(errors),
            "errors": errors[:50]  # Limit to first 50 errors
        }
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to process file: {str(e)}")


@router.post("/bulk-80g-certificates", response_model=dict)
def generate_bulk_80g_certificates(
    donation_ids: List[int] = Query(...),
    financial_year: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Generate 80G certificates for multiple donations in batch
    Returns PDF with all certificates
    """
    temple_id = current_user.temple_id if current_user else None
    
    # Get donations
    donations = db.query(Donation).filter(
        Donation.id.in_(donation_ids),
        Donation.is_cancelled == False
    )
    if temple_id:
        donations = donations.filter(Donation.temple_id == temple_id)
    
    donations = donations.all()
    
    if not donations:
        raise HTTPException(status_code=404, detail="No donations found")
    
    # Filter 80G eligible
    eligible_donations = [d for d in donations if d.category and d.category.is_80g_eligible]
    
    if not eligible_donations:
        raise HTTPException(status_code=400, detail="No 80G eligible donations found")
    
    # Get temple info
    temple = None
    if temple_id:
        temple = db.query(Temple).filter(Temple.id == temple_id).first()
    
    # Create PDF with all certificates
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#FF9933'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    elements.append(Paragraph("80G Tax Exemption Certificates", title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Generate certificate for each donation
    for donation in eligible_donations:
        devotee = db.query(Devotee).filter(Devotee.id == donation.devotee_id).first()
        
        # Certificate content
        cert_data = [
            ['Certificate of Donation - Section 80G'],
            [''],
            ['Temple Name:', temple.name if temple else 'Temple'],
            ['Address:', temple.address if temple else ''],
            ['PAN:', temple.pan_number if temple else ''],
            ['80G Registration:', temple.certificate_80g_number if temple else ''],
            [''],
            ['Donor Details:'],
            ['Name:', devotee.name if devotee else 'Anonymous'],
            ['Address:', devotee.address if devotee else ''],
            [''],
            ['Donation Details:'],
            ['Receipt Number:', donation.receipt_number],
            ['Date:', donation.donation_date.strftime('%d-%m-%Y') if hasattr(donation.donation_date, 'strftime') else str(donation.donation_date)],
            ['Amount:', f"₹{donation.amount:,.2f}"],
            ['Category:', donation.category.name if donation.category else ''],
            [''],
            ['This donation is eligible for tax deduction under Section 80G of the Income Tax Act, 1961.'],
            [''],
        ]
        
        table = Table(cert_data, colWidths=[2*inch, 4*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FF9933')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 0.3*inch))
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"80g_certificates_{financial_year or datetime.now().year}.pdf"
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

