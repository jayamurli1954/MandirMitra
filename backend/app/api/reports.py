"""
Reports API Endpoints
Provides detailed reports for donations and sevas
"""

from fastapi import APIRouter, Depends, Query, HTTPException
from sqlalchemy.orm import Session
from sqlalchemy import func, and_, or_
from datetime import date, datetime, timedelta
from typing import List, Optional
from pydantic import BaseModel

from app.core.database import get_db
from app.core.security import get_current_user
from app.models.user import User
from app.models.donation import Donation, DonationCategory
from app.models.seva import SevaBooking, SevaBookingStatus, Seva
from app.models.devotee import Devotee

router = APIRouter(prefix="/api/v1/reports", tags=["reports"])


# ===== CATEGORY-WISE DONATION REPORT =====


class CategoryWiseDonationItem(BaseModel):
    category: str
    count: int
    amount: float


class CategoryWiseDonationResponse(BaseModel):
    from_date: date
    to_date: date
    categories: List[CategoryWiseDonationItem]
    total_amount: float
    total_count: int


@router.get("/donations/category-wise", response_model=CategoryWiseDonationResponse)
def get_category_wise_donation_report(
    from_date: Optional[date] = Query(None, description="Start date (default: today)"),
    to_date: Optional[date] = Query(None, description="End date (default: today)"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Get category-wise donation report
    Default: Today's donations grouped by category
    Custom: Date range donations grouped by category
    """
    temple_id = current_user.temple_id

    # Default to today if not specified
    if not from_date:
        from_date = date.today()
    if not to_date:
        to_date = date.today()

    # Get donations in date range
    donations = (
        db.query(Donation)
        .filter(
            Donation.temple_id == temple_id,
            Donation.donation_date >= from_date,
            Donation.donation_date <= to_date,
            Donation.is_cancelled == False,
        )
        .all()
    )

    # Group by category
    category_stats = {}
    total_amount = 0.0
    total_count = 0

    for donation in donations:
        category_name = donation.category.name if donation.category else "Unknown"

        if category_name not in category_stats:
            category_stats[category_name] = {"count": 0, "amount": 0.0}

        category_stats[category_name]["count"] += 1
        category_stats[category_name]["amount"] += donation.amount
        total_amount += donation.amount
        total_count += 1

    # Convert to response format
    categories = [
        CategoryWiseDonationItem(category=cat_name, count=stats["count"], amount=stats["amount"])
        for cat_name, stats in sorted(category_stats.items())
    ]

    return CategoryWiseDonationResponse(
        from_date=from_date,
        to_date=to_date,
        categories=categories,
        total_amount=total_amount,
        total_count=total_count,
    )


# ===== DETAILED DONATION REPORT =====


class DetailedDonationItem(BaseModel):
    id: int
    date: date
    receipt_number: str
    devotee_name: str
    devotee_mobile: Optional[str]
    category: str
    payment_mode: str
    amount: float


class DetailedDonationResponse(BaseModel):
    from_date: date
    to_date: date
    donations: List[DetailedDonationItem]
    total_amount: float
    total_count: int


@router.get("/donations/detailed", response_model=DetailedDonationResponse)
def get_detailed_donation_report(
    from_date: date = Query(..., description="Start date"),
    to_date: date = Query(..., description="End date"),
    category: Optional[str] = Query(None, description="Filter by category"),
    payment_mode: Optional[str] = Query(None, description="Filter by payment mode"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Get detailed donation report with:
    - Date, Devotee Name, Mobile Number, Category, Payment Mode, Amount
    """
    # For standalone mode, handle temple_id = None
    temple_id = current_user.temple_id

    # Build query
    donation_filter = [
        Donation.donation_date >= from_date,
        Donation.donation_date <= to_date,
        Donation.is_cancelled == False,
    ]
    if temple_id is not None:
        donation_filter.append(Donation.temple_id == temple_id)

    # Use outerjoin to include anonymous donations (donations without devotee)
    query = db.query(Donation).outerjoin(Devotee).filter(*donation_filter)

    # Apply filters
    if category:
        query = query.join(DonationCategory).filter(DonationCategory.name == category)

    if payment_mode:
        query = query.filter(Donation.payment_mode.ilike(f"%{payment_mode}%"))

    donations = query.order_by(Donation.donation_date.desc(), Donation.id.desc()).all()

    # Convert to response format
    donation_items = []
    total_amount = 0.0

    for donation in donations:
        donation_items.append(
            DetailedDonationItem(
                id=donation.id,
                date=donation.donation_date,
                receipt_number=donation.receipt_number,
                devotee_name=donation.devotee.name if donation.devotee else "Anonymous",
                devotee_mobile=donation.devotee.phone if donation.devotee else None,
                category=donation.category.name if donation.category else "Unknown",
                payment_mode=donation.payment_mode or "N/A",
                amount=donation.amount,
            )
        )
        total_amount += donation.amount

    return DetailedDonationResponse(
        from_date=from_date,
        to_date=to_date,
        donations=donation_items,
        total_amount=total_amount,
        total_count=len(donation_items),
    )


# ===== DETAILED SEVA REPORT =====


class DetailedSevaItem(BaseModel):
    id: int
    receipt_date: date  # Date when money was received (created_at)
    booking_date: date  # Date when seva will be performed
    seva_date: date  # Same as booking_date for now, can be different if rescheduled
    receipt_number: str
    seva_name: str
    devotee_name: str
    devotee_mobile: Optional[str]
    amount: float
    status: str  # Completed or Pending (based on booking_date)
    original_date: Optional[date]  # If rescheduled
    reschedule_reason: Optional[str]


class DetailedSevaResponse(BaseModel):
    from_date: date
    to_date: date
    sevas: List[DetailedSevaItem]
    total_amount: float
    total_count: int
    completed_count: int
    pending_count: int


@router.get("/sevas/detailed", response_model=DetailedSevaResponse)
def get_detailed_seva_report(
    from_date: date = Query(..., description="Start date"),
    to_date: date = Query(..., description="End date"),
    status: Optional[str] = Query(None, description="Filter by status"),
    seva_id: Optional[int] = Query(None, description="Filter by seva ID"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Get detailed seva report with:
    - Receipt Date (when money was received), Seva Date (when seva will be performed)
    - Seva Name, Devotee, Amount, Status
    - Status: Completed (booking_date < today), Pending (booking_date >= today)

    IMPORTANT BUSINESS LOGIC:
    - Sevas for TODAY remain as PENDING until the next day
    - Only sevas from PREVIOUS days are marked as COMPLETED
    - This allows rescheduling of today's bookings if needed
    - Advance bookings can be rescheduled until the seva date passes

    IMPORTANT: Report filters by receipt date (created_at), not booking date.
    This ensures cash/bank reconciliation matches correctly.
    """
    temple_id = current_user.temple_id
    today = date.today()

    # Build query - use with_entities to only select columns that exist
    # This avoids errors if reschedule columns don't exist in database yet
    from sqlalchemy.orm import load_only
    from sqlalchemy import func as sql_func

    # CRITICAL: Filter by receipt date (created_at.date()), not booking_date
    # This ensures reports match cash/bank records and accounting entries
    # Receipt date = when money was received (created_at)
    # Booking date = when seva will be performed (booking_date)

    # Select only the columns we need (excluding reschedule fields that may not exist)
    query = (
        db.query(SevaBooking)
        .options(
            load_only(
                SevaBooking.id,
                SevaBooking.seva_id,
                SevaBooking.devotee_id,
                SevaBooking.booking_date,
                SevaBooking.booking_time,
                SevaBooking.status,
                SevaBooking.amount_paid,
                SevaBooking.receipt_number,
                SevaBooking.special_request,
                SevaBooking.created_at,
            )
        )
        .join(Seva)
        .join(Devotee)
        .filter(
            # Filter by receipt date (created_at.date()), not booking_date
            sql_func.date(SevaBooking.created_at) >= from_date,
            sql_func.date(SevaBooking.created_at) <= to_date,
            SevaBooking.status != SevaBookingStatus.CANCELLED,
        )
    )

    # Apply temple filter - filter through devotee's temple_id
    if temple_id is not None:
        query = query.filter(Devotee.temple_id == temple_id)

    # Apply status filter (status is determined by booking_date, but we filter by receipt date above)
    # IMPORTANT: Today's bookings are PENDING (not completed) to allow rescheduling
    if status:
        if status.lower() == "completed":
            query = query.filter(SevaBooking.booking_date < today)  # Only previous days
        elif status.lower() == "pending":
            query = query.filter(SevaBooking.booking_date >= today)  # Today and future
        else:
            query = query.filter(SevaBooking.status == status)

    # Apply seva filter
    if seva_id:
        query = query.filter(SevaBooking.seva_id == seva_id)

    # Order by receipt date (when money was received), then booking date, then ID
    bookings = query.order_by(
        sql_func.date(SevaBooking.created_at).asc(),
        SevaBooking.booking_date.asc(),
        SevaBooking.id.asc(),
    ).all()

    # Convert to response format
    seva_items = []
    total_amount = 0.0
    completed_count = 0
    pending_count = 0

    for booking in bookings:
        # Get receipt date (when money was received)
        receipt_date = (
            booking.created_at.date() if hasattr(booking.created_at, "date") else booking.created_at
        )

        # Determine status based on booking_date (when seva will be performed)
        # IMPORTANT: Today's bookings remain PENDING until next day to allow rescheduling
        if booking.booking_date < today:
            # Only previous days are marked as completed
            booking_status = "Completed"
            completed_count += 1
        else:
            # Today and future dates are pending (allows rescheduling)
            booking_status = "Pending"
            pending_count += 1

        seva_items.append(
            DetailedSevaItem(
                id=booking.id,
                receipt_date=receipt_date,  # Receipt date (when money was received)
                booking_date=booking.booking_date,  # Booking date (when seva will be performed)
                seva_date=booking.booking_date,  # Can be updated if rescheduled
                receipt_number=booking.receipt_number or f"SEV-{booking.id}",
                seva_name=booking.seva.name_english if booking.seva else "Unknown Seva",
                devotee_name=booking.devotee.name if booking.devotee else "Unknown",
                devotee_mobile=booking.devotee.phone if booking.devotee else None,
                amount=booking.amount_paid,
                status=booking_status,  # Completed or Pending (based on booking_date)
                original_date=None,  # Will be populated when reschedule is implemented
                reschedule_reason=None,
            )
        )
        total_amount += booking.amount_paid

    return DetailedSevaResponse(
        from_date=from_date,
        to_date=to_date,
        sevas=seva_items,
        total_amount=total_amount,
        total_count=len(seva_items),
        completed_count=completed_count,
        pending_count=pending_count,
    )


# ===== 3-DAY SEVA SCHEDULE =====


class SevaScheduleItem(BaseModel):
    id: int
    date: date
    time: Optional[str]
    seva_name: str
    devotee_name: str
    devotee_mobile: Optional[str]
    amount: float
    status: str
    special_request: Optional[str]


class SevaScheduleResponse(BaseModel):
    from_date: date
    to_date: date
    schedule: List[SevaScheduleItem]
    total_bookings: int


@router.get("/sevas/schedule", response_model=SevaScheduleResponse)
def get_seva_schedule(
    days: int = Query(3, ge=1, le=30, description="Number of days ahead (default: 3)"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Get seva schedule for next N days (default: 3 days)
    Helps temple authority make arrangements
    """
    today = date.today()
    end_date = today + timedelta(days=days)

    # Get bookings for next N days - use load_only to avoid reschedule column errors
    from sqlalchemy.orm import load_only

    bookings = (
        db.query(SevaBooking)
        .options(
            load_only(
                SevaBooking.id,
                SevaBooking.seva_id,
                SevaBooking.devotee_id,
                SevaBooking.booking_date,
                SevaBooking.booking_time,
                SevaBooking.status,
                SevaBooking.amount_paid,
                SevaBooking.receipt_number,
                SevaBooking.special_request,
            )
        )
        .join(Seva)
        .join(Devotee)
        .filter(
            SevaBooking.booking_date >= today,
            SevaBooking.booking_date <= end_date,
            SevaBooking.status != SevaBookingStatus.CANCELLED,
        )
        .order_by(SevaBooking.booking_date.asc(), SevaBooking.booking_time.asc())
        .all()
    )

    schedule_items = []

    for booking in bookings:
        # Determine status
        if booking.booking_date == today:
            booking_status = "Today"
        elif booking.booking_date < today:
            booking_status = "Completed"
        else:
            booking_status = "Upcoming"

        schedule_items.append(
            SevaScheduleItem(
                id=booking.id,
                date=booking.booking_date,
                time=booking.booking_time,
                seva_name=booking.seva.name_english if booking.seva else "Unknown Seva",
                devotee_name=booking.devotee.name if booking.devotee else "Unknown",
                devotee_mobile=booking.devotee.phone if booking.devotee else None,
                amount=booking.amount_paid,
                status=booking_status,
                special_request=booking.special_request,
            )
        )

    return SevaScheduleResponse(
        from_date=today,
        to_date=end_date,
        schedule=schedule_items,
        total_bookings=len(schedule_items),
    )


# ==========================================
# EXPORT ENDPOINTS
# ==========================================

from fastapi.responses import StreamingResponse
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_RIGHT


@router.get("/sevas/detailed/export/excel")
def export_detailed_seva_excel(
    from_date: date = Query(...),
    to_date: date = Query(...),
    status: Optional[str] = Query(None),
    seva_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Export Detailed Seva Report to Excel"""
    data = get_detailed_seva_report(from_date, to_date, status, seva_id, db, current_user)

    wb = Workbook()
    ws = wb.active
    ws.title = "Seva Report"

    ws.merge_cells("A1:H1")
    ws["A1"] = f"DETAILED SEVA REPORT ({from_date} to {to_date})"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = [
        "Receipt Date",
        "Seva Date",
        "Receipt No",
        "Seva Name",
        "Devotee",
        "Mobile",
        "Amount",
        "Status",
    ]
    row = 3
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.border = Border(bottom=Side(style="thin"))
    row += 1

    for item in data.sevas:
        ws.cell(
            row=row, column=1, value=item.receipt_date
        )  # Receipt date (when money was received)
        ws.cell(
            row=row, column=2, value=item.booking_date
        )  # Seva date (when seva will be performed)
        ws.cell(row=row, column=3, value=item.receipt_number)
        ws.cell(row=row, column=4, value=item.seva_name)
        ws.cell(row=row, column=5, value=item.devotee_name)
        ws.cell(row=row, column=6, value=item.devotee_mobile or "-")
        ws.cell(row=row, column=7, value=item.amount)
        ws.cell(row=row, column=8, value=item.status)
        row += 1

    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "Total Amount"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"A{row}"].alignment = Alignment(horizontal="right")
    ws[f"G{row}"] = data.total_amount
    ws[f"G{row}"].font = Font(bold=True)

    # Column width adjustment
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 25

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=SevaReport_{from_date}_{to_date}.xlsx"
        },
    )


@router.get("/sevas/detailed/export/pdf")
def export_detailed_seva_pdf(
    from_date: date = Query(...),
    to_date: date = Query(...),
    status: Optional[str] = Query(None),
    seva_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Export Detailed Seva Report to PDF"""
    data = get_detailed_seva_report(from_date, to_date, status, seva_id, db, current_user)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    elements = []
    styles = getSampleStyleSheet()

    title = Paragraph(f"Seva Report ({from_date} to {to_date})", styles["Title"])
    elements.append(title)
    elements.append(Spacer(1, 0.2 * inch))

    table_data = [
        ["Receipt Date", "Seva Date", "Receipt", "Seva Name", "Devotee", "Amount", "Status"]
    ]
    for item in data.sevas:
        table_data.append(
            [
                str(item.receipt_date),  # Receipt date (when money was received)
                str(item.booking_date),  # Seva date (when seva will be performed)
                item.receipt_number,
                Paragraph(item.seva_name[:30], styles["Normal"]),
                Paragraph(item.devotee_name[:25], styles["Normal"]),
                f"{item.amount:,.2f}",
                item.status,
            ]
        )

    table_data.append(["", "", "", "", "Total", f"{data.total_amount:,.2f}", ""])

    t = Table(
        table_data,
        colWidths=[1.2 * inch, 1.2 * inch, 1.5 * inch, 3 * inch, 2 * inch, 1.2 * inch, 1.2 * inch],
    )
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (4, 0), (4, -1), "RIGHT"),
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ]
        )
    )
    elements.append(t)

    doc.build(elements)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename=SevaReport_{from_date}_{to_date}.pdf"
        },
    )
