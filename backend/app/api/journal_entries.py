"""
Journal Entry API Endpoints
Manage double-entry bookkeeping transactions
"""

from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
import io
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from sqlalchemy.orm import Session
from sqlalchemy import func, and_, or_
from typing import List, Optional
from datetime import datetime, date

from app.core.database import get_db
from app.core.security import get_current_user
from app.models.user import User
from app.models.accounting import (
    Account, JournalEntry, JournalLine,
    JournalEntryStatus, AccountType, AccountSubType
)
from app.schemas.accounting import (
    JournalEntryCreate, JournalEntryUpdate, JournalEntryResponse,
    JournalEntryPost, JournalEntryCancel,
    TrialBalanceResponse, TrialBalanceItem,
    AccountLedgerResponse, LedgerEntry,
    ProfitLossResponse, PLCategoryGroup, PLAccountItem,
    CategoryIncomeResponse, CategoryIncomeItem,
    TopDonorsResponse, TopDonorItem,
    BalanceSheetResponse, BalanceSheetGroup, BalanceSheetAccountItem,
    DayBookResponse, DayBookEntry,
    CashBookResponse, CashBookEntry,
    BankBookResponse, BankBookEntry
)
from app.models.donation import Donation
from app.models.devotee import Devotee

router = APIRouter(prefix="/api/v1/journal-entries", tags=["journal-entries"])


# ===== HELPER FUNCTIONS =====

def generate_entry_number(db: Session, temple_id: int) -> str:
    """
    Generate unique journal entry number
    Format: JE/YYYY/0001
    """
    year = datetime.now().year
    prefix = f"JE/{year}/"

    # Get last entry number for this year
    last_entry = db.query(JournalEntry).filter(
        JournalEntry.temple_id == temple_id,
        JournalEntry.entry_number.like(f"{prefix}%")
    ).order_by(JournalEntry.id.desc()).first()

    if last_entry:
        # Extract number and increment
        last_num = int(last_entry.entry_number.split('/')[-1])
        new_num = last_num + 1
    else:
        new_num = 1

    return f"{prefix}{new_num:04d}"


def validate_journal_entry(journal_lines: List, db: Session, temple_id: int):
    """
    Validate journal entry:
    - Must have at least 2 lines
    - Total debits must equal total credits
    - All accounts must exist and belong to temple
    - Each line must have either debit or credit (not both)
    """
    if len(journal_lines) < 2:
        raise HTTPException(
            status_code=400,
            detail="Journal entry must have at least 2 lines (debit and credit)"
        )

    total_debit = sum(line.debit_amount for line in journal_lines)
    total_credit = sum(line.credit_amount for line in journal_lines)

    if abs(total_debit - total_credit) > 0.01:  # Allow small floating point difference
        raise HTTPException(
            status_code=400,
            detail=f"Debits ({total_debit}) must equal credits ({total_credit})"
        )

    # Validate each line
    for line in journal_lines:
        # Check account exists
        account = db.query(Account).filter(
            Account.id == line.account_id,
            Account.temple_id == temple_id
        ).first()

        if not account:
            raise HTTPException(
                status_code=404,
                detail=f"Account ID {line.account_id} not found"
            )

        if not account.is_active:
            raise HTTPException(
                status_code=400,
                detail=f"Account '{account.account_name}' is inactive"
            )

        # Check that line has either debit or credit, not both or neither
        if line.debit_amount > 0 and line.credit_amount > 0:
            raise HTTPException(
                status_code=400,
                detail="A journal line cannot have both debit and credit amounts"
            )

        if line.debit_amount == 0 and line.credit_amount == 0:
            raise HTTPException(
                status_code=400,
                detail="A journal line must have either debit or credit amount"
            )

    return total_debit


# ===== JOURNAL ENTRY CRUD =====

@router.get("/", response_model=List[JournalEntryResponse])
def list_journal_entries(
    status_filter: Optional[JournalEntryStatus] = Query(None, alias="status"),
    from_date: Optional[date] = None,
    to_date: Optional[date] = None,
    reference_type: Optional[str] = None,
    limit: int = Query(100, le=1000),
    offset: int = 0,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Get list of journal entries with filters
    """
    query = db.query(JournalEntry).filter(
        JournalEntry.temple_id == current_user.temple_id
    )

    if status_filter:
        query = query.filter(JournalEntry.status == status_filter)

    if from_date:
        query = query.filter(func.date(JournalEntry.entry_date) >= from_date)

    if to_date:
        query = query.filter(func.date(JournalEntry.entry_date) <= to_date)

    if reference_type:
        query = query.filter(JournalEntry.reference_type == reference_type)

    entries = query.order_by(JournalEntry.entry_date.desc()).limit(limit).offset(offset).all()

    # Populate journal lines with account details
    for entry in entries:
        for line in entry.journal_lines:
            if line.account:
                line.account_code = line.account.account_code
                line.account_name = line.account.account_name

    return entries


@router.get("/{entry_id}", response_model=JournalEntryResponse)
def get_journal_entry(
    entry_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Get journal entry by ID
    """
    entry = db.query(JournalEntry).filter(
        JournalEntry.id == entry_id,
        JournalEntry.temple_id == current_user.temple_id
    ).first()

    if not entry:
        raise HTTPException(status_code=404, detail="Journal entry not found")

    # Populate account details
    for line in entry.journal_lines:
        if line.account:
            line.account_code = line.account.account_code
            line.account_name = line.account.account_name

    return entry


@router.post("/", response_model=JournalEntryResponse)
def create_journal_entry(
    entry_data: JournalEntryCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Create new journal entry (in draft status)
    """
    # Verify temple_id matches current user
    if entry_data.temple_id != current_user.temple_id:
        raise HTTPException(status_code=403, detail="Cannot create entry for different temple")

    # Validate journal entry
    total_amount = validate_journal_entry(entry_data.journal_lines, db, current_user.temple_id)

    # Generate entry number
    entry_number = generate_entry_number(db, current_user.temple_id)

    # Create journal entry
    entry = JournalEntry(
        entry_number=entry_number,
        entry_date=entry_data.entry_date,
        narration=entry_data.narration,
        reference_type=entry_data.reference_type,
        reference_id=entry_data.reference_id,
        temple_id=entry_data.temple_id,
        total_amount=total_amount,
        status=JournalEntryStatus.DRAFT,
        created_by=current_user.id
    )
    db.add(entry)
    db.flush()  # Get entry.id

    # Create journal lines
    for line_data in entry_data.journal_lines:
        line = JournalLine(
            journal_entry_id=entry.id,
            account_id=line_data.account_id,
            debit_amount=line_data.debit_amount,
            credit_amount=line_data.credit_amount,
            description=line_data.description
        )
        db.add(line)

    db.commit()
    db.refresh(entry)

    # Populate account details
    for line in entry.journal_lines:
        if line.account:
            line.account_code = line.account.account_code
            line.account_name = line.account.account_name

    return entry


@router.post("/{entry_id}/post", response_model=JournalEntryResponse)
def post_journal_entry(
    entry_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Post a draft journal entry
    Posted entries cannot be modified
    """
    entry = db.query(JournalEntry).filter(
        JournalEntry.id == entry_id,
        JournalEntry.temple_id == current_user.temple_id
    ).first()

    if not entry:
        raise HTTPException(status_code=404, detail="Journal entry not found")

    if entry.status != JournalEntryStatus.DRAFT:
        raise HTTPException(
            status_code=400,
            detail=f"Only draft entries can be posted. Current status: {entry.status}"
        )

    # Revalidate before posting
    validate_journal_entry(entry.journal_lines, db, current_user.temple_id)

    # Post entry
    entry.status = JournalEntryStatus.POSTED
    entry.posted_by = current_user.id
    entry.posted_at = datetime.utcnow()

    db.commit()
    db.refresh(entry)

    # Populate account details
    for line in entry.journal_lines:
        if line.account:
            line.account_code = line.account.account_code
            line.account_name = line.account.account_name

    return entry


@router.post("/{entry_id}/cancel", response_model=JournalEntryResponse)
def cancel_journal_entry(
    entry_id: int,
    cancel_data: JournalEntryCancel,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Cancel a posted journal entry
    Cancellation creates a reversing entry
    """
    if current_user.role != 'admin':
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only admin users can cancel journal entries"
        )

    entry = db.query(JournalEntry).filter(
        JournalEntry.id == entry_id,
        JournalEntry.temple_id == current_user.temple_id
    ).first()

    if not entry:
        raise HTTPException(status_code=404, detail="Journal entry not found")

    if entry.status != JournalEntryStatus.POSTED:
        raise HTTPException(
            status_code=400,
            detail="Only posted entries can be cancelled"
        )

    # Mark as cancelled
    entry.status = JournalEntryStatus.CANCELLED
    entry.cancelled_by = current_user.id
    entry.cancelled_at = datetime.utcnow()
    entry.cancellation_reason = cancel_data.cancellation_reason

    # Create reversing entry
    reversal_number = generate_entry_number(db, current_user.temple_id)
    reversal_entry = JournalEntry(
        entry_number=reversal_number,
        entry_date=datetime.utcnow(),
        narration=f"Reversal of {entry.entry_number}: {cancel_data.cancellation_reason}",
        reference_type=entry.reference_type,
        reference_id=entry.reference_id,
        temple_id=entry.temple_id,
        total_amount=entry.total_amount,
        status=JournalEntryStatus.POSTED,
        created_by=current_user.id,
        posted_by=current_user.id,
        posted_at=datetime.utcnow()
    )
    db.add(reversal_entry)
    db.flush()

    # Create reversed journal lines (swap debit and credit)
    for line in entry.journal_lines:
        reversed_line = JournalLine(
            journal_entry_id=reversal_entry.id,
            account_id=line.account_id,
            debit_amount=line.credit_amount,  # Swap
            credit_amount=line.debit_amount,  # Swap
            description=f"Reversal: {line.description or ''}"
        )
        db.add(reversed_line)

    db.commit()
    db.refresh(entry)

    return entry


# ===== REPORTS =====

@router.get("/reports/trial-balance", response_model=TrialBalanceResponse)
def get_trial_balance(
    as_of_date: date = Query(default_factory=date.today),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Generate Trial Balance
    Shows accounts with their debit and credit balances
    
    IMPORTANT: Aggregates child accounts under parent accounts for cleaner reporting:
    - Donation Income: Shows 4100 (parent) with sum of all child accounts (4101, 4102, etc.)
    - Seva Income: Shows 4200 (parent) with sum of all child accounts
    - In-Kind: Shows 4400 (parent) with sum of all child accounts
    - Other parent accounts similarly aggregated
    """
    # For standalone mode, handle temple_id = None
    temple_id = current_user.temple_id
    
    # Get all active accounts for temple
    account_filter = [Account.is_active == True]
    if temple_id is not None:
        account_filter.append(Account.temple_id == temple_id)
    
    all_accounts = db.query(Account).filter(*account_filter).order_by(Account.account_code).all()
    
    # Build account map for quick lookup
    account_map = {acc.id: acc for acc in all_accounts}
    
    # Calculate balances for all accounts
    account_balances = {}
    for account in all_accounts:
        balance_filter = [
            JournalLine.account_id == account.id,
            JournalEntry.status == JournalEntryStatus.POSTED,
            func.date(JournalEntry.entry_date) <= as_of_date
        ]
        if temple_id is not None:
            balance_filter.append(JournalEntry.temple_id == temple_id)
        
        balance_query = db.query(
            func.sum(JournalLine.debit_amount).label('total_debit'),
            func.sum(JournalLine.credit_amount).label('total_credit')
        ).join(JournalLine.journal_entry).filter(*balance_filter).first()

        debit = float(balance_query.total_debit or 0) + account.opening_balance_debit
        credit = float(balance_query.total_credit or 0) + account.opening_balance_credit

        net_balance = debit - credit

        # Determine if balance is debit or credit
        if net_balance > 0:
            debit_balance = net_balance
            credit_balance = 0.0
        else:
            debit_balance = 0.0
            credit_balance = abs(net_balance)
        
        account_balances[account.id] = {
            'debit': debit_balance,
            'credit': credit_balance,
            'account': account
        }
    
    # Define parent accounts that should aggregate their children
    # Format: parent_code: [list of child code patterns or ranges]
    parent_accounts_to_aggregate = {
        '4100': ['4101', '4102', '4103', '4104', '4110', '4111', '4112', '4113', '4114', '4115'],  # Donation Income
        '4200': ['4201', '4202', '4203', '4204', '4205', '4206', '4207', '4208', '4209'],  # Seva Income
        '4400': ['4401', '4402', '4403'],  # In-Kind Donation Income
        '4300': ['4301', '4302'],  # Sponsorship Income
    }
    
    # Build list of accounts to show (parent accounts with aggregated children)
    trial_balance_items = []
    total_debits = 0.0
    total_credits = 0.0
    processed_account_ids = set()
    
    # Process parent accounts first (aggregate children)
    for parent_code, child_patterns in parent_accounts_to_aggregate.items():
        parent_account = next((acc for acc in all_accounts if acc.account_code == parent_code), None)
        if not parent_account:
            continue
        
        # Get all child accounts by matching patterns
        child_accounts = []
        for pattern in child_patterns:
            # Find accounts that match the pattern (exact code match)
            matching_accounts = [acc for acc in all_accounts 
                                if acc.account_code == pattern and acc.account_code != parent_code]
            child_accounts.extend(matching_accounts)
        
        # Also find accounts by parent relationship (if parent_account_id is set)
        direct_children = [acc for acc in all_accounts 
                          if acc.parent_account_id == parent_account.id]
        child_accounts.extend(direct_children)
        
        # Remove duplicates
        child_accounts = list({acc.id: acc for acc in child_accounts}.values())
        
        # Aggregate balances: parent + all children
        parent_balance = account_balances.get(parent_account.id, {'debit': 0, 'credit': 0})
        aggregated_debit = parent_balance['debit']
        aggregated_credit = parent_balance['credit']
        
        for child_account in child_accounts:
            if child_account.id in account_balances:
                child_balance = account_balances[child_account.id]
                aggregated_debit += child_balance['debit']
                aggregated_credit += child_balance['credit']
                processed_account_ids.add(child_account.id)
        
        # Only show parent if it has balance (including children)
        if aggregated_debit > 0 or aggregated_credit > 0:
            trial_balance_items.append(TrialBalanceItem(
                account_code=parent_account.account_code,
                account_name=parent_account.account_name,
                account_type=parent_account.account_type,
                debit_balance=aggregated_debit,
                credit_balance=aggregated_credit
            ))
            total_debits += aggregated_debit
            total_credits += aggregated_credit
            processed_account_ids.add(parent_account.id)
    
    # Add all other accounts (not part of aggregated parents)
    for account in all_accounts:
        if account.id in processed_account_ids:
            continue  # Skip already processed accounts
        
        balance = account_balances.get(account.id, {'debit': 0, 'credit': 0})
        if balance['debit'] > 0 or balance['credit'] > 0:
            trial_balance_items.append(TrialBalanceItem(
                account_code=account.account_code,
                account_name=account.account_name,
                account_type=account.account_type,
                debit_balance=balance['debit'],
                credit_balance=balance['credit']
            ))
            total_debits += balance['debit']
            total_credits += balance['credit']

    # Sort by account code
    trial_balance_items.sort(key=lambda x: x.account_code)

    # Check if balanced
    difference = abs(total_debits - total_credits)
    is_balanced = difference < 0.01  # Allow small floating point difference

    return TrialBalanceResponse(
        as_of_date=as_of_date,
        accounts=trial_balance_items,
        total_debits=total_debits,
        total_credits=total_credits,
        is_balanced=is_balanced,
        difference=difference
    )


@router.get("/reports/ledger/{account_id}", response_model=AccountLedgerResponse)
def get_account_ledger(
    account_id: int,
    from_date: date = Query(...),
    to_date: date = Query(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Generate Account Ledger (Statement of Account)
    Shows all transactions for a specific account
    """
    # For standalone mode, handle temple_id = None
    temple_id = current_user.temple_id
    
    account_filter = [Account.id == account_id]
    if temple_id is not None:
        account_filter.append(Account.temple_id == temple_id)
    
    account = db.query(Account).filter(*account_filter).first()

    if not account:
        raise HTTPException(status_code=404, detail="Account not found")

    # Calculate opening balance (as of from_date)
    opening_filter = [
        JournalLine.account_id == account_id,
        JournalEntry.status == JournalEntryStatus.POSTED,
        func.date(JournalEntry.entry_date) < from_date
    ]
    if temple_id is not None:
        opening_filter.append(JournalEntry.temple_id == temple_id)
    
    opening_query = db.query(
        func.sum(JournalLine.debit_amount).label('debit'),
        func.sum(JournalLine.credit_amount).label('credit')
    ).join(JournalLine.journal_entry).filter(*opening_filter).first()

    opening_debit = float(opening_query.debit or 0) + account.opening_balance_debit
    opening_credit = float(opening_query.credit or 0) + account.opening_balance_credit
    opening_balance = opening_debit - opening_credit

    # Get all transactions in date range
    lines_filter = [
        JournalLine.account_id == account_id,
        JournalEntry.status == JournalEntryStatus.POSTED,
        func.date(JournalEntry.entry_date) >= from_date,
        func.date(JournalEntry.entry_date) <= to_date
    ]
    if temple_id is not None:
        lines_filter.append(JournalEntry.temple_id == temple_id)
    
    lines = db.query(JournalLine).join(JournalLine.journal_entry).filter(*lines_filter).order_by(JournalEntry.entry_date).all()

    # Build ledger entries
    ledger_entries = []
    running_balance = opening_balance

    for line in lines:
        running_balance += line.debit_amount - line.credit_amount

        ledger_entries.append(LedgerEntry(
            entry_date=line.journal_entry.entry_date,
            entry_number=line.journal_entry.entry_number,
            narration=line.journal_entry.narration,
            debit_amount=line.debit_amount,
            credit_amount=line.credit_amount,
            running_balance=running_balance,
            reference_type=line.journal_entry.reference_type,
            reference_id=line.journal_entry.reference_id
        ))

    return AccountLedgerResponse(
        account_id=account.id,
        account_code=account.account_code,
        account_name=account.account_name,
        account_type=account.account_type,
        from_date=from_date,
        to_date=to_date,
        opening_balance=opening_balance,
        closing_balance=running_balance,
        entries=ledger_entries
    )


@router.get("/reports/profit-loss", response_model=ProfitLossResponse)
def get_profit_loss_statement(
    from_date: date = Query(...),
    to_date: date = Query(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Generate Profit & Loss Statement (Income & Expenditure Account)
    Shows categorized income and expenses with net surplus/deficit
    """
    # For standalone mode, handle temple_id = None
    temple_id = current_user.temple_id
    
    # Get all income accounts (4000-4999) with balances
    income_filter = [
        Account.account_code.between('4000', '4999'),
        JournalEntry.status == JournalEntryStatus.POSTED,
        func.date(JournalEntry.entry_date) >= from_date,
        func.date(JournalEntry.entry_date) <= to_date
    ]
    if temple_id is not None:
        income_filter.append(Account.temple_id == temple_id)
        income_filter.append(JournalEntry.temple_id == temple_id)
    
    income_accounts = db.query(
        Account.account_code,
        Account.account_name,
        func.sum(JournalLine.credit_amount - JournalLine.debit_amount).label('amount')
    ).join(JournalLine).join(JournalEntry).filter(*income_filter).group_by(Account.account_code, Account.account_name).all()

    # Get all expense accounts (5000-5999) with balances
    expense_filter = [
        Account.account_code.between('5000', '5999'),
        JournalEntry.status == JournalEntryStatus.POSTED,
        func.date(JournalEntry.entry_date) >= from_date,
        func.date(JournalEntry.entry_date) <= to_date
    ]
    if temple_id is not None:
        expense_filter.append(Account.temple_id == temple_id)
        expense_filter.append(JournalEntry.temple_id == temple_id)

    expense_accounts = db.query(
        Account.account_code,
        Account.account_name,
        func.sum(JournalLine.debit_amount - JournalLine.credit_amount).label('amount')
    ).join(JournalLine).join(JournalEntry).filter(*expense_filter).group_by(Account.account_code, Account.account_name).all()

    # Group income by categories
    income_groups = []

    # Donation Income (4100-4199)
    donation_accounts = [acc for acc in income_accounts if '4100' <= acc.account_code <= '4199' and acc.amount > 0]
    if donation_accounts:
        income_groups.append(PLCategoryGroup(
            category_name="Donation Income",
            accounts=[PLAccountItem(account_code=acc.account_code, account_name=acc.account_name, amount=float(acc.amount))
                     for acc in donation_accounts],
            total=sum(float(acc.amount) for acc in donation_accounts)
        ))

    # Seva Income (4200-4299)
    seva_accounts = [acc for acc in income_accounts if '4200' <= acc.account_code <= '4299' and acc.amount > 0]
    if seva_accounts:
        income_groups.append(PLCategoryGroup(
            category_name="Seva Income",
            accounts=[PLAccountItem(account_code=acc.account_code, account_name=acc.account_name, amount=float(acc.amount))
                     for acc in seva_accounts],
            total=sum(float(acc.amount) for acc in seva_accounts)
        ))

    # Sponsorship Income (4300-4399)
    sponsorship_accounts = [acc for acc in income_accounts if '4300' <= acc.account_code <= '4399' and acc.amount > 0]
    if sponsorship_accounts:
        income_groups.append(PLCategoryGroup(
            category_name="Sponsorship Income",
            accounts=[PLAccountItem(account_code=acc.account_code, account_name=acc.account_name, amount=float(acc.amount))
                     for acc in sponsorship_accounts],
            total=sum(float(acc.amount) for acc in sponsorship_accounts)
        ))

    # Other Income (4500-4599)
    other_income_accounts = [acc for acc in income_accounts if '4500' <= acc.account_code <= '4599' and acc.amount > 0]
    if other_income_accounts:
        income_groups.append(PLCategoryGroup(
            category_name="Other Income",
            accounts=[PLAccountItem(account_code=acc.account_code, account_name=acc.account_name, amount=float(acc.amount))
                     for acc in other_income_accounts],
            total=sum(float(acc.amount) for acc in other_income_accounts)
        ))

    total_income = sum(group.total for group in income_groups)

    # Group expenses by categories
    expense_groups = []

    # Operating Expenses (5100-5199)
    operating_accounts = [acc for acc in expense_accounts if '5100' <= acc.account_code <= '5199' and acc.amount > 0]
    if operating_accounts:
        expense_groups.append(PLCategoryGroup(
            category_name="Operating Expenses",
            accounts=[PLAccountItem(account_code=acc.account_code, account_name=acc.account_name, amount=float(acc.amount))
                     for acc in operating_accounts],
            total=sum(float(acc.amount) for acc in operating_accounts)
        ))

    # Administrative Expenses (5200-5299)
    admin_accounts = [acc for acc in expense_accounts if '5200' <= acc.account_code <= '5299' and acc.amount > 0]
    if admin_accounts:
        expense_groups.append(PLCategoryGroup(
            category_name="Administrative Expenses",
            accounts=[PLAccountItem(account_code=acc.account_code, account_name=acc.account_name, amount=float(acc.amount))
                     for acc in admin_accounts],
            total=sum(float(acc.amount) for acc in admin_accounts)
        ))

    # Other Expenses (5300-5999)
    other_expense_accounts = [acc for acc in expense_accounts if '5300' <= acc.account_code <= '5999' and acc.amount > 0]
    if other_expense_accounts:
        expense_groups.append(PLCategoryGroup(
            category_name="Other Expenses",
            accounts=[PLAccountItem(account_code=acc.account_code, account_name=acc.account_name, amount=float(acc.amount))
                     for acc in other_expense_accounts],
            total=sum(float(acc.amount) for acc in other_expense_accounts)
        ))

    total_expenses = sum(group.total for group in expense_groups)
    net_surplus = total_income - total_expenses

    return ProfitLossResponse(
        from_date=from_date,
        to_date=to_date,
        income_groups=income_groups,
        total_income=total_income,
        expense_groups=expense_groups,
        total_expenses=total_expenses,
        net_surplus=net_surplus
    )


@router.get("/reports/category-income", response_model=CategoryIncomeResponse)
def get_category_income_report(
    from_date: date = Query(...),
    to_date: date = Query(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Generate Category-wise Income Report
    Breaks down income by donation categories, seva types, etc.
    """
    # For standalone mode, handle temple_id = None
    temple_id = current_user.temple_id
    
    # Get donation income (4100-4199)
    donation_filter = [
        Account.account_code.between('4100', '4199'),
        JournalEntry.status == JournalEntryStatus.POSTED,
        func.date(JournalEntry.entry_date) >= from_date,
        func.date(JournalEntry.entry_date) <= to_date
    ]
    if temple_id is not None:
        donation_filter.append(Account.temple_id == temple_id)
        donation_filter.append(JournalEntry.temple_id == temple_id)
    
    donation_data = db.query(
        Account.account_code,
        Account.account_name,
        func.sum(JournalLine.credit_amount - JournalLine.debit_amount).label('amount'),
        func.count(JournalLine.id).label('transaction_count')
    ).join(JournalLine).join(JournalEntry).filter(*donation_filter).group_by(Account.account_code, Account.account_name).all()

    # Get seva income (4200-4299)
    seva_filter = [
        Account.account_code.between('4200', '4299'),
        JournalEntry.status == JournalEntryStatus.POSTED,
        func.date(JournalEntry.entry_date) >= from_date,
        func.date(JournalEntry.entry_date) <= to_date
    ]
    if temple_id is not None:
        seva_filter.append(Account.temple_id == temple_id)
        seva_filter.append(JournalEntry.temple_id == temple_id)

    seva_data = db.query(
        Account.account_code,
        Account.account_name,
        func.sum(JournalLine.credit_amount - JournalLine.debit_amount).label('amount'),
        func.count(JournalLine.id).label('transaction_count')
    ).join(JournalLine).join(JournalEntry).filter(*seva_filter).group_by(Account.account_code, Account.account_name).all()

    # Get other income (4300-4599)
    other_filter = [
        Account.account_code.between('4300', '4599'),
        JournalEntry.status == JournalEntryStatus.POSTED,
        func.date(JournalEntry.entry_date) >= from_date,
        func.date(JournalEntry.entry_date) <= to_date
    ]
    if temple_id is not None:
        other_filter.append(Account.temple_id == temple_id)
        other_filter.append(JournalEntry.temple_id == temple_id)

    other_data = db.query(
        Account.account_code,
        Account.account_name,
        func.sum(JournalLine.credit_amount - JournalLine.debit_amount).label('amount'),
        func.count(JournalLine.id).label('transaction_count')
    ).join(JournalLine).join(JournalEntry).filter(*other_filter).group_by(Account.account_code, Account.account_name).all()

    # Calculate total income for percentage
    total_income = sum(float(d.amount) for d in donation_data if d.amount > 0) + \
                  sum(float(s.amount) for s in seva_data if s.amount > 0) + \
                  sum(float(o.amount) for o in other_data if o.amount > 0)

    # Build donation income items
    donation_income = []
    for item in donation_data:
        if item.amount > 0:
            donation_income.append(CategoryIncomeItem(
                account_code=item.account_code,
                account_name=item.account_name,
                amount=float(item.amount),
                percentage=round((float(item.amount) / total_income * 100) if total_income > 0 else 0, 2),
                transaction_count=int(item.transaction_count)
            ))
    donation_income.sort(key=lambda x: x.amount, reverse=True)

    # Build seva income items
    seva_income = []
    for item in seva_data:
        if item.amount > 0:
            seva_income.append(CategoryIncomeItem(
                account_code=item.account_code,
                account_name=item.account_name,
                amount=float(item.amount),
                percentage=round((float(item.amount) / total_income * 100) if total_income > 0 else 0, 2),
                transaction_count=int(item.transaction_count)
            ))
    seva_income.sort(key=lambda x: x.amount, reverse=True)

    # Build other income items
    other_income = []
    for item in other_data:
        if item.amount > 0:
            other_income.append(CategoryIncomeItem(
                account_code=item.account_code,
                account_name=item.account_name,
                amount=float(item.amount),
                percentage=round((float(item.amount) / total_income * 100) if total_income > 0 else 0, 2),
                transaction_count=int(item.transaction_count)
            ))
    other_income.sort(key=lambda x: x.amount, reverse=True)

    return CategoryIncomeResponse(
        from_date=from_date,
        to_date=to_date,
        donation_income=donation_income,
        seva_income=seva_income,
        other_income=other_income,
        total_income=total_income
    )


@router.get("/reports/top-donors", response_model=TopDonorsResponse)
def get_top_donors_report(
    from_date: date = Query(...),
    to_date: date = Query(...),
    limit: int = Query(10, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Generate Top Donors Report
    Shows top donors by total donation amount
    """
    # For standalone mode, handle temple_id = None
    temple_id = current_user.temple_id
    
    # Get donations with devotee info
    donor_filter = [
        func.date(Donation.donation_date) >= from_date,
        func.date(Donation.donation_date) <= to_date,
        Donation.is_cancelled == False
    ]
    if temple_id is not None:
        donor_filter.append(Devotee.temple_id == temple_id)
        donor_filter.append(Donation.temple_id == temple_id)
    
    donations = db.query(
        Devotee.id.label('devotee_id'),
        Devotee.name.label('devotee_name'),
        func.sum(Donation.amount).label('total_donated'),
        func.count(Donation.id).label('donation_count'),
        func.max(Donation.donation_date).label('last_donation_date')
    ).join(Donation, Devotee.id == Donation.devotee_id).filter(*donor_filter).group_by(Devotee.id, Devotee.name).order_by(
        func.sum(Donation.amount).desc()
    ).limit(limit).all()

    donors = []
    total_amount = 0.0

    for donor in donations:
        # Get categories this donor donated to
        categories_query = db.query(
            func.distinct(Account.account_name)
        ).join(JournalLine).join(JournalEntry).join(Donation, JournalEntry.reference_id == Donation.id).filter(
            Donation.devotee_id == donor.devotee_id,
            JournalEntry.reference_type == TransactionType.DONATION,
            Account.account_code.between('4100', '4199'),
            func.date(Donation.donation_date) >= from_date,
            func.date(Donation.donation_date) <= to_date
        ).all()

        categories = [cat[0] for cat in categories_query] if categories_query else []

        donors.append(TopDonorItem(
            devotee_id=donor.devotee_id,
            devotee_name=donor.devotee_name,
            total_donated=float(donor.total_donated),
            donation_count=int(donor.donation_count),
            last_donation_date=donor.last_donation_date.date() if hasattr(donor.last_donation_date, 'date') else donor.last_donation_date,
            categories=categories
        ))

        total_amount += float(donor.total_donated)

    return TopDonorsResponse(
        from_date=from_date,
        to_date=to_date,
        donors=donors,
        total_donors=len(donors),
        total_amount=total_amount
    )


@router.get("/reports/balance-sheet", response_model=BalanceSheetResponse)
def get_balance_sheet(
    as_of_date: date = Query(default_factory=date.today),
    include_previous_year: bool = Query(False, description="Include previous year comparison"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Generate Balance Sheet - Financial Position Statement
    Shows Assets, Liabilities, and Funds as of a specific date
    Format: Schedule III compliant (adapted for trusts)
    """
    # For standalone mode, handle temple_id = None
    temple_id = current_user.temple_id
    
    # Calculate previous year date if needed
    previous_year_date = None
    if include_previous_year:
        from datetime import timedelta
        # Approximate: 1 year before
        previous_year_date = date(as_of_date.year - 1, as_of_date.month, as_of_date.day)
    
    # Helper function to calculate account balance
    def get_account_balance(account_id: int, as_of: date) -> float:
        """Calculate account balance as of date"""
        balance_filter = [
            JournalLine.account_id == account_id,
            JournalEntry.status == JournalEntryStatus.POSTED,
            func.date(JournalEntry.entry_date) <= as_of
        ]
        if temple_id is not None:
            balance_filter.append(JournalEntry.temple_id == temple_id)
        
        balance_query = db.query(
            func.sum(JournalLine.debit_amount).label('total_debit'),
            func.sum(JournalLine.credit_amount).label('total_credit')
        ).join(JournalLine.journal_entry).filter(*balance_filter).first()
        
        account = db.query(Account).filter(Account.id == account_id).first()
        if not account:
            return 0.0
        
        debit = float(balance_query.total_debit or 0) + account.opening_balance_debit
        credit = float(balance_query.total_credit or 0) + account.opening_balance_credit
        
        # For assets: debit balance is positive, credit balance is negative
        # For liabilities: credit balance is positive, debit balance is negative
        if account.account_type == AccountType.ASSET:
            return debit - credit
        elif account.account_type in [AccountType.LIABILITY, AccountType.EQUITY]:
            return credit - debit
        else:
            return debit - credit
    
    # Get all accounts
    account_filter = [Account.is_active == True]
    if temple_id is not None:
        account_filter.append(Account.temple_id == temple_id)
    
    all_accounts = db.query(Account).filter(*account_filter).order_by(Account.account_code).all()
    
    # ===== ASSETS SIDE =====
    
    # Fixed Assets (1000-1999, account_type = ASSET, account_subtype = FIXED_ASSET)
    fixed_asset_accounts = [
        acc for acc in all_accounts
        if acc.account_type == AccountType.ASSET
        and acc.account_subtype == AccountSubType.FIXED_ASSET
        and acc.account_code.startswith(('1', '2'))
    ]
    
    fixed_asset_groups = []
    fixed_assets_total = 0.0
    prev_fixed_assets_total = 0.0
    
    # Group fixed assets by category
    fixed_asset_categories = {}
    for acc in fixed_asset_accounts:
        # Determine category from account code or name
        category = "Other Fixed Assets"
        if 'land' in acc.account_name.lower() or 'building' in acc.account_name.lower():
            category = "Land & Buildings"
        elif 'vehicle' in acc.account_name.lower():
            category = "Vehicles"
        elif 'furniture' in acc.account_name.lower() or 'fixture' in acc.account_name.lower():
            category = "Furniture & Fixtures"
        elif 'equipment' in acc.account_name.lower() or 'computer' in acc.account_name.lower():
            category = "Equipment"
        
        if category not in fixed_asset_categories:
            fixed_asset_categories[category] = []
        
        current_bal = get_account_balance(acc.id, as_of_date)
        prev_bal = get_account_balance(acc.id, previous_year_date) if previous_year_date else None
        
        if abs(current_bal) > 0.01:  # Only include non-zero balances
            fixed_asset_categories[category].append(BalanceSheetAccountItem(
                account_code=acc.account_code,
                account_name=acc.account_name,
                current_year=current_bal,
                previous_year=prev_bal
            ))
            fixed_assets_total += current_bal
            if prev_bal:
                prev_fixed_assets_total += prev_bal
    
    for category, accounts in fixed_asset_categories.items():
        if accounts:
            group_total = sum(acc.current_year for acc in accounts)
            prev_total = sum(acc.previous_year or 0 for acc in accounts) if include_previous_year else None
            fixed_asset_groups.append(BalanceSheetGroup(
                group_name=category,
                accounts=accounts,
                group_total=group_total,
                previous_year_total=prev_total
            ))
    
    # Current Assets (1000-1999, account_type = ASSET, account_subtype = CURRENT_ASSET or CASH_BANK)
    current_asset_accounts = [
        acc for acc in all_accounts
        if acc.account_type == AccountType.ASSET
        and acc.account_subtype in [AccountSubType.CURRENT_ASSET, AccountSubType.CASH_BANK, AccountSubType.RECEIVABLE]
        and acc.account_code.startswith(('1', '2'))
    ]
    
    current_asset_groups = []
    current_assets_total = 0.0
    prev_current_assets_total = 0.0
    
    # Group current assets
    current_asset_categories = {
        "Cash & Bank": [],
        "Investments": [],
        "Loans & Advances": [],
        "Other Current Assets": []
    }
    
    for acc in current_asset_accounts:
        category = "Other Current Assets"
        if 'cash' in acc.account_name.lower() or 'bank' in acc.account_name.lower():
            category = "Cash & Bank"
        elif 'investment' in acc.account_name.lower() or 'deposit' in acc.account_name.lower():
            category = "Investments"
        elif 'advance' in acc.account_name.lower() or 'loan' in acc.account_name.lower():
            category = "Loans & Advances"
        
        current_bal = get_account_balance(acc.id, as_of_date)
        prev_bal = get_account_balance(acc.id, previous_year_date) if previous_year_date else None
        
        if abs(current_bal) > 0.01:
            current_asset_categories[category].append(BalanceSheetAccountItem(
                account_code=acc.account_code,
                account_name=acc.account_name,
                current_year=current_bal,
                previous_year=prev_bal
            ))
            current_assets_total += current_bal
            if prev_bal:
                prev_current_assets_total += prev_bal
    
    for category, accounts in current_asset_categories.items():
        if accounts:
            group_total = sum(acc.current_year for acc in accounts)
            prev_total = sum(acc.previous_year or 0 for acc in accounts) if include_previous_year else None
            current_asset_groups.append(BalanceSheetGroup(
                group_name=category,
                accounts=accounts,
                group_total=group_total,
                previous_year_total=prev_total
            ))
    
    total_assets = fixed_assets_total + current_assets_total
    prev_total_assets = (prev_fixed_assets_total + prev_current_assets_total) if include_previous_year else None
    
    # ===== LIABILITIES & FUNDS SIDE =====
    
    # Corpus Fund / Capital Fund (3000-3999, account_type = EQUITY, account_subtype = CORPUS_FUND)
    corpus_accounts = [
        acc for acc in all_accounts
        if acc.account_type == AccountType.EQUITY
        and acc.account_subtype == AccountSubType.CORPUS_FUND
        and acc.account_code.startswith('3')
    ]
    
    corpus_fund = 0.0
    prev_corpus_fund = 0.0
    for acc in corpus_accounts:
        corpus_fund += get_account_balance(acc.id, as_of_date)
        if previous_year_date:
            prev_corpus_fund += get_account_balance(acc.id, previous_year_date)
    
    # Designated Funds (3000-3999, account_type = EQUITY, account_subtype = GENERAL_FUND or others)
    designated_fund_accounts = [
        acc for acc in all_accounts
        if acc.account_type == AccountType.EQUITY
        and acc.account_subtype == AccountSubType.GENERAL_FUND
        and acc.account_code.startswith('3')
    ]
    
    designated_fund_groups = []
    for acc in designated_fund_accounts:
        current_bal = get_account_balance(acc.id, as_of_date)
        prev_bal = get_account_balance(acc.id, previous_year_date) if previous_year_date else None
        
        if abs(current_bal) > 0.01:
            designated_fund_groups.append(BalanceSheetGroup(
                group_name=acc.account_name,
                accounts=[BalanceSheetAccountItem(
                    account_code=acc.account_code,
                    account_name=acc.account_name,
                    current_year=current_bal,
                    previous_year=prev_bal
                )],
                group_total=current_bal,
                previous_year_total=prev_bal
            ))
    
    # Current Liabilities (2000-2999, account_type = LIABILITY)
    current_liability_accounts = [
        acc for acc in all_accounts
        if acc.account_type == AccountType.LIABILITY
        and acc.account_subtype == AccountSubType.CURRENT_LIABILITY
        and acc.account_code.startswith('2')
    ]
    
    current_liability_groups = []
    current_liabilities_total = 0.0
    prev_current_liabilities_total = 0.0
    
    liability_categories = {
        "Sundry Creditors": [],
        "Expenses Payable": [],
        "Advance from Devotees": [],
        "TDS Payable": [],
        "Other Liabilities": []
    }
    
    for acc in current_liability_accounts:
        category = "Other Liabilities"
        if 'creditor' in acc.account_name.lower():
            category = "Sundry Creditors"
        elif 'payable' in acc.account_name.lower() and 'tds' not in acc.account_name.lower():
            category = "Expenses Payable"
        elif 'advance' in acc.account_name.lower() or 'devotee' in acc.account_name.lower():
            category = "Advance from Devotees"
        elif 'tds' in acc.account_name.lower():
            category = "TDS Payable"
        
        current_bal = get_account_balance(acc.id, as_of_date)
        prev_bal = get_account_balance(acc.id, previous_year_date) if previous_year_date else None
        
        if abs(current_bal) > 0.01:
            liability_categories[category].append(BalanceSheetAccountItem(
                account_code=acc.account_code,
                account_name=acc.account_name,
                current_year=current_bal,
                previous_year=prev_bal
            ))
            current_liabilities_total += current_bal
            if prev_bal:
                prev_current_liabilities_total += prev_bal
    
    for category, accounts in liability_categories.items():
        if accounts:
            group_total = sum(acc.current_year for acc in accounts)
            prev_total = sum(acc.previous_year or 0 for acc in accounts) if include_previous_year else None
            current_liability_groups.append(BalanceSheetGroup(
                group_name=category,
                accounts=accounts,
                group_total=group_total,
                previous_year_total=prev_total
            ))
    
    # Calculate total liabilities and funds
    total_designated_funds = sum(group.group_total for group in designated_fund_groups)
    total_liabilities_and_funds = corpus_fund + total_designated_funds + current_liabilities_total
    prev_total_liabilities = None
    if include_previous_year:
        prev_designated = sum(group.previous_year_total or 0 for group in designated_fund_groups)
        prev_total_liabilities = prev_corpus_fund + prev_designated + prev_current_liabilities_total
    
    # Check if balanced
    difference = abs(total_assets - total_liabilities_and_funds)
    is_balanced = difference < 0.01
    
    return BalanceSheetResponse(
        as_of_date=as_of_date,
        previous_year_date=previous_year_date,
        fixed_assets=fixed_asset_groups,
        current_assets=current_asset_groups,
        total_assets=total_assets,
        previous_year_total_assets=prev_total_assets,
        corpus_fund=corpus_fund,
        previous_year_corpus_fund=prev_corpus_fund if include_previous_year else None,
        designated_funds=designated_fund_groups,
        current_liabilities=current_liability_groups,
        total_liabilities_and_funds=total_liabilities_and_funds,
        previous_year_total_liabilities=prev_total_liabilities,
        is_balanced=is_balanced,
        difference=difference
    )


@router.get("/reports/day-book", response_model=DayBookResponse)
def get_day_book(
    date: date = Query(default_factory=date.today, description="Date for day book"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Generate Day Book - All transactions for a specific day
    Shows all receipts and payments for the day with opening/closing balance
    """
    temple_id = current_user.temple_id
    
    # Get all journal entries for the day
    entry_filter = [
        func.date(JournalEntry.entry_date) == date,
        JournalEntry.status == JournalEntryStatus.POSTED
    ]
    if temple_id is not None:
        entry_filter.append(JournalEntry.temple_id == temple_id)
    
    entries = db.query(JournalEntry).filter(*entry_filter).order_by(JournalEntry.entry_date).all()
    
    # Calculate opening balance (balance before this date)
    opening_filter = [
        func.date(JournalEntry.entry_date) < date,
        JournalEntry.status == JournalEntryStatus.POSTED
    ]
    if temple_id is not None:
        opening_filter.append(JournalEntry.temple_id == temple_id)
    
    # Get cash and bank accounts for opening balance
    cash_accounts = db.query(Account).filter(
        Account.account_type == AccountType.ASSET,
        Account.account_subtype.in_([AccountSubType.CASH_BANK, AccountSubType.CURRENT_ASSET]),
        Account.is_active == True
    )
    if temple_id is not None:
        cash_accounts = cash_accounts.filter(Account.temple_id == temple_id)
    cash_accounts = cash_accounts.all()
    
    opening_balance = 0.0
    for acc in cash_accounts:
        balance_filter = [
            JournalLine.account_id == acc.id,
            JournalEntry.status == JournalEntryStatus.POSTED,
            func.date(JournalEntry.entry_date) < date
        ]
        if temple_id is not None:
            balance_filter.append(JournalEntry.temple_id == temple_id)
        
        balance_query = db.query(
            func.sum(JournalLine.debit_amount).label('total_debit'),
            func.sum(JournalLine.credit_amount).label('total_credit')
        ).join(JournalLine.journal_entry).filter(*balance_filter).first()
        
        debit = float(balance_query.total_debit or 0) + acc.opening_balance_debit
        credit = float(balance_query.total_credit or 0) + acc.opening_balance_credit
        opening_balance += (debit - credit)
    
    # Separate receipts and payments
    receipts = []
    payments = []
    total_receipts = 0.0
    total_payments = 0.0
    
    for entry in entries:
        # Get all lines for this entry
        lines = db.query(JournalLine).filter(JournalLine.journal_entry_id == entry.id).all()
        
        for line in lines:
            account = db.query(Account).filter(Account.id == line.account_id).first()
            if not account:
                continue
            
            # Determine if this is a receipt (money coming in) or payment (money going out)
            is_cash_bank = account.account_subtype in [AccountSubType.CASH_BANK, AccountSubType.CURRENT_ASSET]
            is_income = account.account_type == AccountType.INCOME
            is_expense = account.account_type == AccountType.EXPENSE
            
            if is_cash_bank and line.debit_amount > 0:
                # Money received (cash/bank debited)
                receipts.append(DayBookEntry(
                    entry_number=entry.entry_number,
                    entry_date=entry.entry_date,
                    narration=entry.narration or line.description or "",
                    voucher_type=entry.reference_type or "Manual",
                    debit_amount=line.debit_amount,
                    credit_amount=0.0,
                    account_name=account.account_name,
                    party_name=None
                ))
                total_receipts += line.debit_amount
            elif is_cash_bank and line.credit_amount > 0:
                # Money paid (cash/bank credited)
                payments.append(DayBookEntry(
                    entry_number=entry.entry_number,
                    entry_date=entry.entry_date,
                    narration=entry.narration or line.description or "",
                    voucher_type=entry.reference_type or "Manual",
                    debit_amount=0.0,
                    credit_amount=line.credit_amount,
                    account_name=account.account_name,
                    party_name=None
                ))
                total_payments += line.credit_amount
            elif is_income and line.credit_amount > 0:
                # Income recognized (receipt)
                receipts.append(DayBookEntry(
                    entry_number=entry.entry_number,
                    entry_date=entry.entry_date,
                    narration=entry.narration or line.description or "",
                    voucher_type=entry.reference_type or "Manual",
                    debit_amount=0.0,
                    credit_amount=line.credit_amount,
                    account_name=account.account_name,
                    party_name=None
                ))
                total_receipts += line.credit_amount
            elif is_expense and line.debit_amount > 0:
                # Expense incurred (payment)
                payments.append(DayBookEntry(
                    entry_number=entry.entry_number,
                    entry_date=entry.entry_date,
                    narration=entry.narration or line.description or "",
                    voucher_type=entry.reference_type or "Manual",
                    debit_amount=line.debit_amount,
                    credit_amount=0.0,
                    account_name=account.account_name,
                    party_name=None
                ))
                total_payments += line.debit_amount
    
    net_cash_flow = total_receipts - total_payments
    closing_balance = opening_balance + net_cash_flow
    
    return DayBookResponse(
        date=date,
        opening_balance=opening_balance,
        receipts=receipts,
        total_receipts=total_receipts,
        payments=payments,
        total_payments=total_payments,
        net_cash_flow=net_cash_flow,
        closing_balance=closing_balance
    )


@router.get("/reports/cash-book", response_model=CashBookResponse)
def get_cash_book(
    from_date: date = Query(...),
    to_date: date = Query(...),
    counter_id: Optional[int] = Query(None, description="Counter ID if multiple counters"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Generate Cash Book - All cash transactions for a date range
    Shows all cash receipts and payments with running balance
    """
    temple_id = current_user.temple_id
    
    # Get cash accounts
    cash_account_filter = [
        Account.account_type == AccountType.ASSET,
        Account.account_subtype == AccountSubType.CASH_BANK,
        Account.is_active == True
    ]
    if temple_id is not None:
        cash_account_filter.append(Account.temple_id == temple_id)
    if counter_id:
        # If counter-specific, filter by account code pattern
        cash_account_filter.append(Account.account_code.like(f'110{counter_id}%'))
    
    cash_accounts = db.query(Account).filter(*cash_account_filter).all()
    
    if not cash_accounts:
        # Return empty cash book
        return CashBookResponse(
            from_date=from_date,
            to_date=to_date,
            opening_balance=0.0,
            entries=[],
            closing_balance=0.0,
            total_receipts=0.0,
            total_payments=0.0
        )
    
    cash_account_ids = [acc.id for acc in cash_accounts]
    
    # Calculate opening balance
    opening_filter = [
        JournalLine.account_id.in_(cash_account_ids),
        JournalEntry.status == JournalEntryStatus.POSTED,
        func.date(JournalEntry.entry_date) < from_date
    ]
    if temple_id is not None:
        opening_filter.append(JournalEntry.temple_id == temple_id)
    
    opening_query = db.query(
        func.sum(JournalLine.debit_amount).label('total_debit'),
        func.sum(JournalLine.credit_amount).label('total_credit')
    ).join(JournalLine.journal_entry).filter(*opening_filter).first()
    
    opening_debit = float(opening_query.total_debit or 0)
    opening_credit = float(opening_query.total_credit or 0)
    
    # Add opening balances from accounts
    for acc in cash_accounts:
        opening_debit += acc.opening_balance_debit
        opening_credit += acc.opening_balance_credit
    
    opening_balance = opening_debit - opening_credit
    
    # Get all transactions in date range
    lines_filter = [
        JournalLine.account_id.in_(cash_account_ids),
        JournalEntry.status == JournalEntryStatus.POSTED,
        func.date(JournalEntry.entry_date) >= from_date,
        func.date(JournalEntry.entry_date) <= to_date
    ]
    if temple_id is not None:
        lines_filter.append(JournalEntry.temple_id == temple_id)
    
    lines = db.query(JournalLine).join(JournalLine.journal_entry).filter(*lines_filter).order_by(
        JournalEntry.entry_date, JournalEntry.id
    ).all()
    
    # Build cash book entries
    entries = []
    running_balance = opening_balance
    total_receipts = 0.0
    total_payments = 0.0
    
    for line in lines:
        account = db.query(Account).filter(Account.id == line.account_id).first()
        entry = line.journal_entry
        
        receipt_amount = 0.0
        payment_amount = 0.0
        
        if line.debit_amount > 0:
            # Cash received
            receipt_amount = line.debit_amount
            total_receipts += receipt_amount
            running_balance += receipt_amount
        elif line.credit_amount > 0:
            # Cash paid
            payment_amount = line.credit_amount
            total_payments += payment_amount
            running_balance -= payment_amount
        
        entries.append(CashBookEntry(
            date=entry.entry_date.date() if hasattr(entry.entry_date, 'date') else entry.entry_date,
            entry_number=entry.entry_number,
            narration=entry.narration or line.description or "",
            receipt_amount=receipt_amount,
            payment_amount=payment_amount,
            running_balance=running_balance,
            voucher_type=entry.reference_type or "Manual",
            party_name=None
        ))
    
    return CashBookResponse(
        from_date=from_date,
        to_date=to_date,
        opening_balance=opening_balance,
        entries=entries,
        closing_balance=running_balance,
        total_receipts=total_receipts,
        total_payments=total_payments
    )


@router.get("/reports/bank-book", response_model=BankBookResponse)
def get_bank_book(
    account_id: int = Query(..., description="Bank account ID"),
    from_date: date = Query(...),
    to_date: date = Query(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Generate Bank Book - All bank transactions for a specific bank account
    Shows deposits, withdrawals, and cheque tracking
    """
    temple_id = current_user.temple_id
    
    # Get bank account
    account_filter = [Account.id == account_id]
    if temple_id is not None:
        account_filter.append(Account.temple_id == temple_id)
    
    account = db.query(Account).filter(*account_filter).first()
    
    if not account:
        raise HTTPException(status_code=404, detail="Bank account not found")
    
    if account.account_subtype != AccountSubType.CASH_BANK:
        raise HTTPException(status_code=400, detail="Account is not a bank account")
    
    # Calculate opening balance
    opening_filter = [
        JournalLine.account_id == account_id,
        JournalEntry.status == JournalEntryStatus.POSTED,
        func.date(JournalEntry.entry_date) < from_date
    ]
    if temple_id is not None:
        opening_filter.append(JournalEntry.temple_id == temple_id)
    
    opening_query = db.query(
        func.sum(JournalLine.debit_amount).label('total_debit'),
        func.sum(JournalLine.credit_amount).label('total_credit')
    ).join(JournalLine.journal_entry).filter(*opening_filter).first()
    
    opening_debit = float(opening_query.total_debit or 0) + account.opening_balance_debit
    opening_credit = float(opening_query.total_credit or 0) + account.opening_balance_credit
    opening_balance = opening_debit - opening_credit
    
    # Get all transactions in date range
    lines_filter = [
        JournalLine.account_id == account_id,
        JournalEntry.status == JournalEntryStatus.POSTED,
        func.date(JournalEntry.entry_date) >= from_date,
        func.date(JournalEntry.entry_date) <= to_date
    ]
    if temple_id is not None:
        lines_filter.append(JournalEntry.temple_id == temple_id)
    
    lines = db.query(JournalLine).join(JournalLine.journal_entry).filter(*lines_filter).order_by(
        JournalEntry.entry_date, JournalEntry.id
    ).all()
    
    # Build bank book entries
    entries = []
    outstanding_cheques = []
    running_balance = opening_balance
    total_deposits = 0.0
    total_withdrawals = 0.0
    
    for line in lines:
        entry = line.journal_entry
        
        deposit_amount = 0.0
        withdrawal_amount = 0.0
        cheque_number = None
        cleared = True
        
        if line.debit_amount > 0:
            # Deposit
            deposit_amount = line.debit_amount
            total_deposits += deposit_amount
            running_balance += deposit_amount
        elif line.credit_amount > 0:
            # Withdrawal
            withdrawal_amount = line.credit_amount
            total_withdrawals += withdrawal_amount
            running_balance -= withdrawal_amount
            
            # Try to extract cheque number from narration or reference
            narration = entry.narration or line.description or ""
            if 'cheque' in narration.lower() or 'chq' in narration.lower():
                # Extract cheque number (simple pattern matching)
                chq_match = re.search(r'ch[eq]*\s*[#:]?\s*(\d+)', narration, re.IGNORECASE)
                if chq_match:
                    cheque_number = chq_match.group(1)
                    # For now, assume cheques are cleared. In real system, track clearance status
                    cleared = True
        
        entries.append(BankBookEntry(
            date=entry.entry_date.date() if hasattr(entry.entry_date, 'date') else entry.entry_date,
            entry_number=entry.entry_number,
            narration=entry.narration or line.description or "",
            cheque_number=cheque_number,
            deposit_amount=deposit_amount,
            withdrawal_amount=withdrawal_amount,
            running_balance=running_balance,
            voucher_type=entry.reference_type or "Manual",
            cleared=cleared
        ))
        
        # Track outstanding cheques (not cleared)
        if cheque_number and not cleared:
            outstanding_cheques.append(BankBookEntry(
                date=entry.entry_date.date() if hasattr(entry.entry_date, 'date') else entry.entry_date,
                entry_number=entry.entry_number,
                narration=entry.narration or line.description or "",
                cheque_number=cheque_number,
                deposit_amount=0.0,
                withdrawal_amount=withdrawal_amount,
                running_balance=0.0,
                voucher_type=entry.reference_type or "Manual",
                cleared=False
            ))
    
    # Get bank details from account name or separate bank table (if exists)
    bank_name = None
    account_number = None
    # Try to extract from account name (e.g., "Bank - SBI Current Account")
    if ' - ' in account.account_name:
        parts = account.account_name.split(' - ')
        if len(parts) >= 2:
            bank_name = parts[0].replace('Bank', '').strip()
            account_number = parts[1] if len(parts) > 1 else None
    
    return BankBookResponse(
        account_id=account_id,
        account_code=account.account_code,
        account_name=account.account_name,
        bank_name=bank_name,
        account_number=account_number,
        from_date=from_date,
        to_date=to_date,
        opening_balance=opening_balance,
        entries=entries,
        closing_balance=running_balance,
        total_deposits=total_deposits,
        total_withdrawals=total_withdrawals,
        outstanding_cheques=outstanding_cheques
    )


# ==========================================
# EXPORT ENDPOINTS
# ==========================================

@router.get("/reports/day-book/export/excel")
def export_day_book_excel(
    date: date = Query(default_factory=date.today),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Export Day Book to Excel"""
    # Get data using existing logic
    data = get_day_book(date, db, current_user)
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Day Book - {date}"
    
    # Headers
    ws.merge_cells('A1:F1')
    ws['A1'] = "MANDIRSYNC DAY BOOK REPORT"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:F2')
    ws['A2'] = f"Date: {date.strftime('%d-%m-%Y')}"
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Opening Balance
    ws.merge_cells('A4:E4')
    ws['A4'] = "Opening Balance"
    ws['A4'].font = Font(bold=True)
    ws['F4'] = data.opening_balance
    ws['F4'].font = Font(bold=True)
    
    # Receipts Section
    row = 6
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "RECEIPTS"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'A{row}'].alignment = Alignment(horizontal='center')
    row += 1
    
    headers = ["Entry No", "Account", "Narration", "Voucher Type", "Amount"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.border = Border(bottom=Side(style='thin'))
    row += 1
    
    for entry in data.receipts:
        ws.cell(row=row, column=1, value=entry.entry_number)
        ws.cell(row=row, column=2, value=entry.account_name)
        ws.cell(row=row, column=3, value=entry.narration)
        ws.cell(row=row, column=4, value=entry.voucher_type)
        amount = entry.debit_amount if entry.debit_amount > 0 else entry.credit_amount
        ws.cell(row=row, column=5, value=amount)
        row += 1
        
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "Total Receipts"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'A{row}'].alignment = Alignment(horizontal='right')
    ws[f'E{row}'] = data.total_receipts
    ws[f'E{row}'].font = Font(bold=True)
    row += 2
    
    # Payments Section
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "PAYMENTS"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'A{row}'].alignment = Alignment(horizontal='center')
    row += 1
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.border = Border(bottom=Side(style='thin'))
    row += 1
    
    for entry in data.payments:
        ws.cell(row=row, column=1, value=entry.entry_number)
        ws.cell(row=row, column=2, value=entry.account_name)
        ws.cell(row=row, column=3, value=entry.narration)
        ws.cell(row=row, column=4, value=entry.voucher_type)
        amount = entry.credit_amount if entry.credit_amount > 0 else entry.debit_amount
        ws.cell(row=row, column=5, value=amount)
        row += 1
        
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "Total Payments"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'A{row}'].alignment = Alignment(horizontal='right')
    ws[f'E{row}'] = data.total_payments
    ws[f'E{row}'].font = Font(bold=True)
    row += 2
    
    # Closing Balance
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "Closing Balance"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'F{row}'] = data.closing_balance
    ws[f'F{row}'].font = Font(bold=True)
    
    # Adjust column widths
    for col in range(1, 7):
        ws.column_dimensions[chr(64+col)].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 35
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=DayBook_{date}.xlsx"}
    )

@router.get("/reports/day-book/export/pdf")
def export_day_book_pdf(
    date: date = Query(default_factory=date.today),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Export Day Book to PDF"""
    data = get_day_book(date, db, current_user)
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title = Paragraph(f"Day Book Report - {date.strftime('%d-%m-%Y')}", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Opening Balance
    elements.append(Paragraph(f"<b>Opening Balance:</b> ₹{data.opening_balance:,.2f}", styles['Normal']))
    elements.append(Spacer(1, 0.2*inch))
    
    # Receipts Table
    elements.append(Paragraph("<b>RECEIPTS</b>", styles['Heading3']))
    receipt_data = [['Entry No', 'Account', 'Narration', 'Amount']]
    for entry in data.receipts:
        amount = entry.debit_amount if entry.debit_amount > 0 else entry.credit_amount
        receipt_data.append([
            entry.entry_number,
            entry.account_name,
            Paragraph(entry.narration[:50], styles['Normal']),
            f"₹{amount:,.2f}"
        ])
    
    # Add total receipt row
    receipt_data.append(['', '', 'Total Receipts', f"₹{data.total_receipts:,.2f}"])
    
    r_table = Table(receipt_data, colWidths=[1.2*inch, 2*inch, 2.5*inch, 1*inch])
    r_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (-1, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'), # Total row bold
    ]))
    elements.append(r_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Payments Table
    elements.append(Paragraph("<b>PAYMENTS</b>", styles['Heading3']))
    payment_data = [['Entry No', 'Account', 'Narration', 'Amount']]
    for entry in data.payments:
        amount = entry.credit_amount if entry.credit_amount > 0 else entry.debit_amount
        payment_data.append([
            entry.entry_number,
            entry.account_name,
            Paragraph(entry.narration[:50], styles['Normal']),
            f"₹{amount:,.2f}"
        ])
    
    # Add total payment row
    payment_data.append(['', '', 'Total Payments', f"₹{data.total_payments:,.2f}"])
    
    p_table = Table(payment_data, colWidths=[1.2*inch, 2*inch, 2.5*inch, 1*inch])
    p_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (-1, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
    ]))
    elements.append(p_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Closing Balance
    elements.append(Paragraph(f"<b>Closing Balance:</b> ₹{data.closing_balance:,.2f}", styles['Normal']))
    
    doc.build(elements)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=DayBook_{date}.pdf"}
    )

@router.get("/reports/cash-book/export/excel")
def export_cash_book_excel(
    from_date: date = Query(...),
    to_date: date = Query(...),
    counter_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Export Cash Book to Excel"""
    data = get_cash_book(from_date, to_date, counter_id, db, current_user)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Cash Book"
    
    ws.merge_cells('A1:G1')
    ws['A1'] = f"CASH BOOK REPORT ({from_date} to {to_date})"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws['A3'] = f"Opening Balance: {data.opening_balance}"
    ws['A3'].font = Font(bold=True)
    
    headers = ["Date", "Entry No", "Narration", "Receipt", "Payment", "Balance", "Voucher Type"]
    row = 5
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.border = Border(bottom=Side(style='thin'))
    row += 1
    
    for entry in data.entries:
        ws.cell(row=row, column=1, value=entry.date)
        ws.cell(row=row, column=2, value=entry.entry_number)
        ws.cell(row=row, column=3, value=entry.narration)
        ws.cell(row=row, column=4, value=entry.receipt_amount)
        ws.cell(row=row, column=5, value=entry.payment_amount)
        ws.cell(row=row, column=6, value=entry.running_balance)
        ws.cell(row=row, column=7, value=entry.voucher_type)
        row += 1
        
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'] = "Totals"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'A{row}'].alignment = Alignment(horizontal='right')
    ws[f'D{row}'] = data.total_receipts
    ws[f'D{row}'].font = Font(bold=True)
    ws[f'E{row}'] = data.total_payments
    ws[f'E{row}'].font = Font(bold=True)
    row += 1
    
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = f"Closing Balance: {data.closing_balance}"
    ws[f'A{row}'].font = Font(bold=True)
    
    # Columns width
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['C'].width = 40
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=CashBook_{from_date}_{to_date}.xlsx"}
    )

@router.get("/reports/cash-book/export/pdf")
def export_cash_book_pdf(
    from_date: date = Query(...),
    to_date: date = Query(...),
    counter_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Export Cash Book to PDF"""
    data = get_cash_book(from_date, to_date, counter_id, db, current_user)
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    elements = []
    styles = getSampleStyleSheet()
    
    title = Paragraph(f"Cash Book Report ({from_date} to {to_date})", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    elements.append(Paragraph(f"<b>Opening Balance:</b> ₹{data.opening_balance:,.2f}", styles['Normal']))
    elements.append(Spacer(1, 0.2*inch))
    
    table_data = [['Date', 'Entry No', 'Narration', 'Receipt', 'Payment', 'Balance']]
    for entry in data.entries:
        table_data.append([
            str(entry.date),
            entry.entry_number,
            Paragraph(entry.narration[:40], styles['Normal']),
            f"{entry.receipt_amount:,.2f}",
            f"{entry.payment_amount:,.2f}",
            f"{entry.running_balance:,.2f}"
        ])
        
    # Total Row
    table_data.append([
        '', '', 'Totals', 
        f"{data.total_receipts:,.2f}", 
        f"{data.total_payments:,.2f}", 
        f"{data.closing_balance:,.2f}"
    ])
    
    t = Table(table_data, colWidths=[1.2*inch, 1.2*inch, 3*inch, 1.2*inch, 1.2*inch, 1.5*inch])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (3, 0), (-1, -1), 'RIGHT'), # Amounts right aligned
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
    ]))
    elements.append(t)
    
    doc.build(elements)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=CashBook_{from_date}_{to_date}.pdf"}
    )

@router.get("/reports/bank-book/export/excel")
def export_bank_book_excel(
    account_id: int = Query(...),
    from_date: date = Query(...),
    to_date: date = Query(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Export Bank Book to Excel"""
    data = get_bank_book(account_id, from_date, to_date, db, current_user)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Bank Book"
    
    ws.merge_cells('A1:G1')
    ws['A1'] = f"BANK BOOK: {data.account_name} ({data.bank_name or 'Bank'})"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws['A2'] = f"Period: {from_date} to {to_date}"
    ws['A3'] = f"Account No: {data.account_number or 'N/A'}"
    ws['A4'] = f"Opening Balance: {data.opening_balance}"
    ws['A4'].font = Font(bold=True)
    
    headers = ["Date", "Entry No", "Narration", "Cheque No", "Deposit", "Withdrawal", "Balance"]
    row = 6
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.border = Border(bottom=Side(style='thin'))
    row += 1
    
    for entry in data.entries:
        ws.cell(row=row, column=1, value=entry.date)
        ws.cell(row=row, column=2, value=entry.entry_number)
        ws.cell(row=row, column=3, value=entry.narration)
        ws.cell(row=row, column=4, value=entry.cheque_number or "-")
        ws.cell(row=row, column=5, value=entry.deposit_amount)
        ws.cell(row=row, column=6, value=entry.withdrawal_amount)
        ws.cell(row=row, column=7, value=entry.running_balance)
        row += 1
    
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "Totals"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'A{row}'].alignment = Alignment(horizontal='right')
    ws[f'E{row}'] = data.total_deposits
    ws[f'E{row}'].font = Font(bold=True)
    ws[f'F{row}'] = data.total_withdrawals
    ws[f'F{row}'].font = Font(bold=True)
    row += 1
    
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = f"Closing Balance: {data.closing_balance}"
    ws[f'A{row}'].font = Font(bold=True)
    
    # Columns width
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['C'].width = 40
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=BankBook_{from_date}_{to_date}.xlsx"}
    )

@router.get("/reports/bank-book/export/pdf")
def export_bank_book_pdf(
    account_id: int = Query(...),
    from_date: date = Query(...),
    to_date: date = Query(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Export Bank Book to PDF"""
    data = get_bank_book(account_id, from_date, to_date, db, current_user)
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    elements = []
    styles = getSampleStyleSheet()
    
    title = Paragraph(f"Bank Book: {data.account_name} ({from_date} to {to_date})", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    elements.append(Paragraph(f"<b>Opening Balance:</b> ₹{data.opening_balance:,.2f}", styles['Normal']))
    elements.append(Spacer(1, 0.2*inch))
    
    table_data = [['Date', 'Narration', 'Cheque', 'Deposit', 'Withdrawal', 'Balance']]
    for entry in data.entries:
        table_data.append([
            str(entry.date),
            Paragraph(entry.narration[:30], styles['Normal']),
            entry.cheque_number or '-',
            f"{entry.deposit_amount:,.2f}",
            f"{entry.withdrawal_amount:,.2f}",
            f"{entry.running_balance:,.2f}"
        ])
        
    table_data.append([
        '', 'Totals', '',
        f"{data.total_deposits:,.2f}",
        f"{data.total_withdrawals:,.2f}",
        f"{data.closing_balance:,.2f}"
    ])
    
    t = Table(table_data, colWidths=[1.2*inch, 3.5*inch, 1*inch, 1.2*inch, 1.2*inch, 1.5*inch])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
    ]))
    elements.append(t)
    
    doc.build(elements)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=BankBook_{from_date}_{to_date}.pdf"}
    )

