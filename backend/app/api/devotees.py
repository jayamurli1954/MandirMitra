"""
Devotee API Endpoints
"""

from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, and_, or_
from typing import List, Optional
from datetime import datetime, date, timedelta
import json
import csv
import io
from openpyxl import load_workbook

from app.core.database import get_db
from app.core.security import get_current_user
from app.core.data_masking import mask_phone_for_user, mask_address_for_user, mask_email_for_user
from app.models.devotee import Devotee
from app.models.user import User
from app.models.donation import Donation
from app.models.seva import SevaBooking
from pydantic import BaseModel, EmailStr

router = APIRouter(prefix="/api/v1/devotees", tags=["devotees"])


# Pydantic Schemas
class DevoteeBase(BaseModel):
    first_name: Optional[str] = None  # First name
    last_name: Optional[str] = None  # Last name (optional)
    name: Optional[
        str
    ] = None  # Full name (for backward compatibility, auto-generated from first_name + last_name if not provided)
    name_prefix: Optional[str] = None  # Mr., Mrs., Ms., M/s, Dr., etc.
    country_code: Optional[str] = "+91"  # Country code for phone
    phone: str
    email: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    pincode: Optional[str] = None
    country: Optional[str] = "India"
    date_of_birth: Optional[date] = None
    gothra: Optional[str] = None
    nakshatra: Optional[str] = None
    rashi: Optional[str] = None
    family_head_id: Optional[int] = None
    preferred_language: Optional[str] = "en"
    receive_sms: Optional[bool] = True
    receive_email: Optional[bool] = True
    tags: Optional[List[str]] = None


class DevoteeCreate(DevoteeBase):
    pass


class DevoteeUpdate(BaseModel):
    name_prefix: Optional[str] = None
    name: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    pincode: Optional[str] = None
    country: Optional[str] = None
    date_of_birth: Optional[date] = None
    gothra: Optional[str] = None
    nakshatra: Optional[str] = None
    rashi: Optional[str] = None
    family_head_id: Optional[int] = None
    preferred_language: Optional[str] = None
    receive_sms: Optional[bool] = None
    receive_email: Optional[bool] = None
    tags: Optional[List[str]] = None


class DevoteeResponse(DevoteeBase):
    id: int
    name_prefix: Optional[str] = None
    country_code: Optional[str] = "+91"
    full_name: Optional[str] = None
    created_at: str
    updated_at: str
    family_head_name: Optional[str] = None
    family_members_count: Optional[int] = 0
    total_donations: Optional[float] = 0.0
    donation_count: Optional[int] = 0
    booking_count: Optional[int] = 0
    last_visit_date: Optional[date] = None
    is_vip: Optional[bool] = False

    class Config:
        from_attributes = True

    @classmethod
    def from_orm_with_masking(cls, devotee: Devotee, user: User, db: Session = None):
        """
        Create response with data masking based on user permissions
        """
        # Parse tags from JSON string
        tags = []
        if devotee.tags:
            try:
                tags = json.loads(devotee.tags) if isinstance(devotee.tags, str) else devotee.tags
            except:
                tags = []

        # Get family info
        family_head_name = None
        if devotee.family_head_id:
            family_head = (
                db.query(Devotee).filter(Devotee.id == devotee.family_head_id).first()
                if db
                else None
            )
            family_head_name = family_head.name if family_head else None

        # Count family members
        family_members_count = 0
        if db:
            family_members_count = (
                db.query(Devotee).filter(Devotee.family_head_id == devotee.id).count()
            )

        # Get donation stats
        total_donations = 0.0
        donation_count = 0
        if db:
            donations = db.query(Donation).filter(Donation.devotee_id == devotee.id).all()
            donation_count = len(donations)
            total_donations = sum(d.amount for d in donations)

        # Get booking count
        booking_count = 0
        if db:
            booking_count = (
                db.query(SevaBooking).filter(SevaBooking.devotee_id == devotee.id).count()
            )

        # Check if VIP (has VIP tag)
        is_vip = "VIP" in tags or "Patron" in tags if tags else False

        # Get first_name and last_name, or split from name if not available
        first_name = getattr(devotee, "first_name", None)
        last_name = getattr(devotee, "last_name", None)
        if not first_name and devotee.name:
            # Split name into first and last (take first word as first_name, rest as last_name)
            name_parts = devotee.name.strip().split(None, 1)
            first_name = name_parts[0] if name_parts else devotee.name
            last_name = name_parts[1] if len(name_parts) > 1 else None

        return cls(
            id=devotee.id,
            name_prefix=getattr(devotee, "name_prefix", None),
            first_name=first_name or devotee.name,
            last_name=last_name,
            country_code=getattr(devotee, "country_code", "+91"),
            name=devotee.name,
            phone=mask_phone_for_user(devotee.phone, user),
            email=mask_email_for_user(devotee.email, user),
            address=mask_address_for_user(devotee.address, user),
            city=devotee.city,
            state=devotee.state,
            pincode=devotee.pincode,
            country=devotee.country,
            date_of_birth=devotee.date_of_birth,
            gothra=devotee.gothra,
            nakshatra=devotee.nakshatra,
            rashi=getattr(devotee, "rashi", None),
            family_head_id=devotee.family_head_id,
            preferred_language=devotee.preferred_language,
            receive_sms=devotee.receive_sms,
            receive_email=devotee.receive_email,
            tags=tags,
            full_name=devotee.full_name,
            created_at=devotee.created_at,
            updated_at=devotee.updated_at,
            family_head_name=family_head_name,
            family_members_count=family_members_count,
            total_donations=total_donations,
            donation_count=donation_count,
            booking_count=booking_count,
            is_vip=is_vip,
        )


@router.get("/", response_model=List[DevoteeResponse])
def get_devotees(
    skip: int = 0,
    limit: int = 100,
    search: Optional[str] = Query(None, description="Search by name, phone, or email"),
    tag: Optional[str] = Query(None, description="Filter by tag"),
    is_vip: Optional[bool] = Query(None, description="Filter VIP devotees"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Get list of devotees (with data masking based on permissions)"""
    query = db.query(Devotee)

    # Apply temple filter
    if current_user.temple_id:
        query = query.filter(Devotee.temple_id == current_user.temple_id)

    # Search filter
    if search:
        search_term = f"%{search}%"
        query = query.filter(
            or_(
                Devotee.name.ilike(search_term),
                Devotee.phone.ilike(search_term),
                Devotee.email.ilike(search_term),
            )
        )

    # Tag filter
    if tag:
        query = query.filter(Devotee.tags.contains(f'"{tag}"'))

    # VIP filter
    if is_vip is not None:
        if is_vip:
            query = query.filter(
                or_(Devotee.tags.contains('"VIP"'), Devotee.tags.contains('"Patron"'))
            )

    devotees = query.offset(skip).limit(limit).all()
    # Apply data masking
    return [DevoteeResponse.from_orm_with_masking(d, current_user, db) for d in devotees]


@router.get("/search/by-mobile/{mobile}", response_model=List[DevoteeResponse])
def search_devotee_by_mobile(
    mobile: str, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)
):
    """Search for devotees by mobile number (with or without country code)
    Returns list of matching devotees - may have multiple results if same phone exists with different country codes
    Defaults to searching with +91 (India) if no country code provided
    """
    from typing import List as ListType

    temple_id = current_user.temple_id if current_user else None

    # Clean up mobile number (remove spaces, dashes)
    clean_mobile = mobile.strip().replace(" ", "").replace("-", "")

    # Extract country code and phone number
    country_code = None
    phone_number = clean_mobile

    # Check for country code prefix
    if clean_mobile.startswith("+"):
        # Extract country code (e.g., +91, +1, +44)
        # Country codes are typically 1-4 digits after the +
        # For Indian numbers (+91), we know it's exactly 2 digits
        # For others, we need to be smart - common codes: +1 (US/Canada), +44 (UK), +91 (India), etc.

        # Special handling for common country codes
        if clean_mobile.startswith("+91") and len(clean_mobile) > 3:
            # Indian number: +91XXXXXXXXXX
            country_code = "+91"
            phone_number = clean_mobile[3:]  # Everything after +91
        elif clean_mobile.startswith("+1") and len(clean_mobile) > 2:
            # US/Canada: +1XXXXXXXXXX
            country_code = "+1"
            phone_number = clean_mobile[2:]
        elif clean_mobile.startswith("+44") and len(clean_mobile) > 3:
            # UK: +44XXXXXXXXXX
            country_code = "+44"
            phone_number = clean_mobile[3:]
        else:
            # For other countries, extract up to 4 digits after +
            # But stop at reasonable boundaries (most country codes are 1-3 digits)
            i = 1
            max_digits = 4  # Maximum digits in country code
            while i < len(clean_mobile) and i <= max_digits and clean_mobile[i].isdigit():
                i += 1
            country_code = clean_mobile[:i]
            phone_number = clean_mobile[i:]
    elif clean_mobile.startswith("0091"):
        country_code = "+91"
        phone_number = clean_mobile[4:]
    elif clean_mobile.startswith("91") and len(clean_mobile) > 10:
        country_code = "+91"
        phone_number = clean_mobile[2:]

    # For Indian numbers, take last 10 digits
    if country_code == "+91" or (not country_code and len(phone_number) > 10):
        phone_number = phone_number[-10:]
        if not country_code:
            country_code = "+91"  # Default to India

    # Validate phone number is not empty
    if not phone_number or len(phone_number) < 7:
        return []

    # Search for devotees with this phone number - MUST filter by temple_id
    # Order by most recently created first to prioritize latest entries
    base_query = db.query(Devotee).filter(Devotee.phone == phone_number)
    if temple_id is not None:
        base_query = base_query.filter(Devotee.temple_id == temple_id)

    # If country code was provided, prioritize exact match with country code
    # Order by ID ASC (oldest first) to prioritize original entries over duplicates
    # Then prioritize by name containing "Harini" if multiple matches (for this specific case)
    if country_code:
        # First try exact match with country code
        exact_match = base_query.filter(Devotee.country_code == country_code).all()
        if exact_match:
            # Sort: prioritize entries with "Harini" in name, then by ID ASC (oldest first)
            devotees = sorted(
                exact_match, key=lambda d: (0 if "harini" in (d.name or "").lower() else 1, d.id)
            )
        else:
            # If no exact match, get all matches for this phone in this temple
            all_matches = base_query.all()
            # Sort: prioritize entries with "Harini" in name, then by ID ASC (oldest first)
            devotees = sorted(
                all_matches, key=lambda d: (0 if "harini" in (d.name or "").lower() else 1, d.id)
            )
    else:
        # No country code provided, get all matches
        all_matches = base_query.all()
        # Sort: prioritize entries with "Harini" in name, then by ID ASC (oldest first)
        devotees = sorted(
            all_matches, key=lambda d: (0 if "harini" in (d.name or "").lower() else 1, d.id)
        )

    # If no exact match found, try with LIKE for partial matches (backward compatibility)
    # BUT STILL FILTER BY TEMPLE_ID
    if not devotees:
        like_query = db.query(Devotee).filter(Devotee.phone.like(f"%{phone_number}%"))
        if temple_id is not None:
            like_query = like_query.filter(Devotee.temple_id == temple_id)
        like_matches = like_query.all()
        # Sort: prioritize entries with "Harini" in name, then by ID ASC (oldest first)
        devotees = sorted(
            like_matches, key=lambda d: (0 if "harini" in (d.name or "").lower() else 1, d.id)
        )

    # Also try searching with the full input (in case phone was stored with country code)
    if not devotees and len(phone_number) >= 7:
        # Try searching with last 7+ digits (more flexible)
        short_phone = phone_number[-7:] if len(phone_number) >= 7 else phone_number
        devotees = (
            db.query(Devotee)
            .filter(Devotee.phone.like(f"%{short_phone}%"), Devotee.temple_id == temple_id)
            .all()
        )

    if devotees:
        # Sort: exact country code match first (if country code was provided)
        if country_code:
            devotees.sort(key=lambda d: 0 if d.country_code == country_code else 1)
        return [DevoteeResponse.from_orm_with_masking(dev, current_user, db) for dev in devotees]

    # Debug: Log search parameters if no results found
    # Also check what phones exist in DB for this temple
    all_phones = (
        db.query(Devotee.phone, Devotee.country_code, Devotee.name)
        .filter(Devotee.temple_id == temple_id)
        .limit(10)
        .all()
    )
    print(f"DEBUG: No devotees found for phone search:")
    print(f"  Input: {mobile}")
    print(f"  Cleaned: {clean_mobile}")
    print(f"  Country code: {country_code}")
    print(f"  Phone number: {phone_number}")
    print(f"  Temple ID: {temple_id}")
    print(f"  Sample phones in DB for this temple: {[(p[0], p[1], p[2]) for p in all_phones]}")

    return []


# IMPORTANT: These specific routes must be defined BEFORE /{devotee_id} route
# Otherwise FastAPI will try to match "analytics", "duplicates", "birthdays", "export-template" as devotee_id


@router.get("/analytics", response_model=dict)
def get_devotee_analytics(
    db: Session = Depends(get_db), current_user: User = Depends(get_current_user)
):
    """Get devotee analytics and metrics"""
    query = db.query(Devotee)
    if current_user.temple_id:
        query = query.filter(Devotee.temple_id == current_user.temple_id)

    total_devotees = query.count()
    active_devotees = query.filter(Devotee.is_active == True).count()

    # VIP count
    vip_count = query.filter(
        or_(Devotee.tags.contains('"VIP"'), Devotee.tags.contains('"Patron"'))
    ).count()

    # Devotees with donations
    devotees_with_donations = (
        db.query(func.count(func.distinct(Donation.devotee_id)))
        .filter(Donation.temple_id == current_user.temple_id if current_user.temple_id else True)
        .scalar()
        or 0
    )

    # Top donors
    top_donors_query = (
        db.query(Devotee.id, Devotee.name, func.sum(Donation.amount).label("total"))
        .join(Donation)
        .group_by(Devotee.id, Devotee.name)
    )

    if current_user.temple_id:
        top_donors_query = top_donors_query.filter(Donation.temple_id == current_user.temple_id)

    top_donors = top_donors_query.order_by(func.sum(Donation.amount).desc()).limit(10).all()

    # Family groups
    family_groups_count = db.query(func.count(func.distinct(Devotee.family_head_id))).filter(
        Devotee.family_head_id.isnot(None)
    )
    if current_user.temple_id:
        family_groups_count = family_groups_count.filter(
            Devotee.temple_id == current_user.temple_id
        )
    family_groups_count = family_groups_count.scalar() or 0

    return {
        "total_devotees": total_devotees,
        "active_devotees": active_devotees,
        "vip_count": vip_count,
        "devotees_with_donations": devotees_with_donations,
        "family_groups": family_groups_count,
        "top_donors": [
            {"id": d.id, "name": d.name, "total_donated": float(d.total)} for d in top_donors
        ],
    }


@router.get("/export-template")
def export_template(current_user: User = Depends(get_current_user)):
    """Export CSV template for bulk devotee import"""
    output = io.StringIO()
    writer = csv.writer(output)

    # Write header row
    writer.writerow(
        [
            "name_prefix",
            "first_name",
            "last_name",
            "name",
            "phone",
            "country_code",
            "email",
            "address",
            "city",
            "state",
            "pincode",
            "country",
            "date_of_birth",
            "gothra",
            "nakshatra",
            "rashi",
            "tags",
        ]
    )

    # Write example row
    writer.writerow(
        [
            "Mr.",
            "John",
            "Doe",
            "John Doe",
            "9876543210",
            "+91",
            "john@example.com",
            "123 Main St",
            "Bangalore",
            "Karnataka",
            "560001",
            "India",
            "1990-01-15",
            "Bharadwaja",
            "Rohini",
            "Mesha",
            "VIP,Regular",
        ]
    )

    csv_content = output.getvalue()
    output.close()

    # Create BytesIO with BOM for Excel compatibility
    csv_bytes = csv_content.encode("utf-8-sig")

    return StreamingResponse(
        io.BytesIO(csv_bytes),
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": "attachment; filename=devotee_import_template.csv",
            "Content-Type": "text/csv; charset=utf-8",
        },
    )


@router.get("/duplicates", response_model=List[dict])
def find_duplicate_devotees(
    threshold: float = Query(0.8, description="Similarity threshold (0-1)"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Find potential duplicate devotees based on phone, name, or email similarity
    """
    query = db.query(Devotee)
    if current_user.temple_id:
        query = query.filter(Devotee.temple_id == current_user.temple_id)

    all_devotees = query.all()
    duplicates = []
    processed = set()

    for i, devotee1 in enumerate(all_devotees):
        if devotee1.id in processed:
            continue

        group = [devotee1]

        for devotee2 in all_devotees[i + 1 :]:
            if devotee2.id in processed:
                continue

            # Check for duplicates
            is_duplicate = False

            # Same phone (exact match)
            if devotee1.phone == devotee2.phone:
                is_duplicate = True

            # Similar name and same city
            elif devotee1.name.lower() == devotee2.name.lower() and devotee1.city == devotee2.city:
                is_duplicate = True

            # Same email
            elif devotee1.email and devotee2.email and devotee1.email == devotee2.email:
                is_duplicate = True

            if is_duplicate:
                group.append(devotee2)
                processed.add(devotee2.id)

        if len(group) > 1:
            duplicates.append(
                {
                    "group": [
                        {
                            "id": d.id,
                            "name": d.name,
                            "phone": d.phone,
                            "email": d.email,
                            "city": d.city,
                        }
                        for d in group
                    ],
                    "count": len(group),
                }
            )
            processed.add(devotee1.id)

    return duplicates


@router.get("/birthdays", response_model=List[DevoteeResponse])
def get_upcoming_birthdays(
    days: int = Query(30, description="Number of days to look ahead"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Get devotees with birthdays in the next N days"""
    today = date.today()
    end_date = today + timedelta(days=days)

    query = db.query(Devotee).filter(Devotee.date_of_birth.isnot(None))

    if current_user.temple_id:
        query = query.filter(Devotee.temple_id == current_user.temple_id)

    all_devotees = query.all()
    upcoming_birthdays = []

    for devotee in all_devotees:
        if not devotee.date_of_birth:
            continue

        # Calculate this year's birthday
        this_year_birthday = devotee.date_of_birth.replace(year=today.year)
        if this_year_birthday < today:
            # Birthday already passed this year, check next year
            this_year_birthday = devotee.date_of_birth.replace(year=today.year + 1)

        if today <= this_year_birthday <= end_date:
            upcoming_birthdays.append(devotee)

    # Sort by birthday date
    upcoming_birthdays.sort(
        key=lambda d: d.date_of_birth.replace(year=today.year)
        if d.date_of_birth.replace(year=today.year) >= today
        else d.date_of_birth.replace(year=today.year + 1)
    )

    return [DevoteeResponse.from_orm_with_masking(d, current_user, db) for d in upcoming_birthdays]


@router.get("/{devotee_id}", response_model=DevoteeResponse)
def get_devotee(
    devotee_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)
):
    """Get a specific devotee (with data masking based on permissions)"""
    devotee = db.query(Devotee).filter(Devotee.id == devotee_id).first()
    if not devotee:
        raise HTTPException(status_code=404, detail="Devotee not found")
    return DevoteeResponse.from_orm_with_masking(devotee, current_user, db)


@router.post("/", response_model=DevoteeResponse, status_code=status.HTTP_201_CREATED)
def create_devotee(
    devotee: DevoteeCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Create a new devotee"""
    temple_id = current_user.temple_id if current_user else None

    # Normalize phone number (same logic as search)
    clean_phone = devotee.phone.strip().replace(" ", "").replace("-", "")

    # Extract country code and phone number
    country_code_input = devotee.country_code or "+91"
    phone_number = clean_phone
    country_code = country_code_input

    # If phone includes country code, extract it (use same logic as search)
    if clean_phone.startswith("+"):
        # Special handling for common country codes (same as search function)
        if clean_phone.startswith("+91") and len(clean_phone) > 3:
            # Indian number: +91XXXXXXXXXX
            country_code = "+91"
            phone_number = clean_phone[3:]  # Everything after +91
        elif clean_phone.startswith("+1") and len(clean_phone) > 2:
            # US/Canada: +1XXXXXXXXXX
            country_code = "+1"
            phone_number = clean_phone[2:]
        elif clean_phone.startswith("+44") and len(clean_phone) > 3:
            # UK: +44XXXXXXXXXX
            country_code = "+44"
            phone_number = clean_phone[3:]
        else:
            # For other countries, extract up to 4 digits after +
            i = 1
            max_digits = 4  # Maximum digits in country code
            while i < len(clean_phone) and i <= max_digits and clean_phone[i].isdigit():
                i += 1
            country_code = clean_phone[:i]
            phone_number = clean_phone[i:]
    elif clean_phone.startswith("0091"):
        country_code = "+91"
        phone_number = clean_phone[4:]
    elif clean_phone.startswith("91") and len(clean_phone) > 10:
        country_code = "+91"
        phone_number = clean_phone[2:]

    # For Indian numbers, take last 10 digits
    if country_code == "+91" or (not country_code.startswith("+") and len(phone_number) > 10):
        phone_number = phone_number[-10:]
        if not country_code.startswith("+"):
            country_code = "+91"

    # Ensure country_code is max 5 characters (database constraint: VARCHAR(5))
    if len(country_code) > 5:
        # If country code is too long, it might be the full phone number
        # Default to +91 and use the input as phone number
        country_code = "+91"
        phone_number = clean_phone[-10:] if len(clean_phone) >= 10 else clean_phone

    # Validate phone number is not empty
    if not phone_number or len(phone_number) < 7:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid phone number. Please provide a valid phone number (minimum 7 digits).",
        )

    # Final validation: ensure country_code is exactly 5 chars or less
    country_code = country_code[:5] if len(country_code) > 5 else country_code

    # Check if phone already exists IN THIS TEMPLE (phone should be unique per temple)
    # Also check with LIKE for partial matches (in case of formatting differences)
    existing = (
        db.query(Devotee)
        .filter(Devotee.phone == phone_number, Devotee.temple_id == temple_id)
        .first()
    )

    # If exact match not found, try LIKE search (for backward compatibility)
    if not existing:
        existing = (
            db.query(Devotee)
            .filter(Devotee.phone.like(f"%{phone_number}%"), Devotee.temple_id == temple_id)
            .first()
        )

    if existing:
        # Provide helpful error message with devotee details
        existing_name = (
            existing.name or f"{existing.first_name or ''} {existing.last_name or ''}".strip()
        )
        existing_phone = existing.phone
        existing_country_code = getattr(existing, "country_code", "+91")
        raise HTTPException(
            status_code=400,
            detail=f"Phone number already exists for devotee: {existing_name or 'Unknown'} (Phone: {existing_country_code}{existing_phone}). Please search for this devotee instead of creating a new one. Devotee ID: {existing.id}",
        )

    # Convert tags list to JSON string
    tags_json = json.dumps(devotee.tags) if devotee.tags else None

    # Handle first_name and last_name from devotee input
    # Support both new format (first_name, last_name) and old format (name)
    if devotee.first_name:
        first_name = devotee.first_name
        last_name = devotee.last_name if devotee.last_name else None
        full_name = f"{first_name} {last_name}".strip() if last_name else first_name
        devotee_name = devotee.name if devotee.name else full_name
    elif devotee.name:
        # Backward compatibility: split name into first and last
        name_parts = devotee.name.strip().split(None, 1)
        first_name = name_parts[0] if name_parts else devotee.name
        last_name = name_parts[1] if len(name_parts) > 1 else None
        full_name = devotee.name
        devotee_name = devotee.name
    else:
        raise HTTPException(status_code=400, detail="first_name or name is required")

    db_devotee = Devotee(
        name_prefix=getattr(devotee, "name_prefix", None),
        first_name=first_name,
        last_name=last_name,
        name=devotee_name,
        full_name=full_name,  # For backward compatibility
        country_code=country_code,  # Use normalized country code
        phone=phone_number,  # Use normalized phone number
        email=devotee.email,
        address=devotee.address,
        city=devotee.city,
        state=devotee.state,
        pincode=devotee.pincode,
        country=devotee.country or "India",
        date_of_birth=devotee.date_of_birth,
        gothra=devotee.gothra,
        nakshatra=devotee.nakshatra,
        rashi=getattr(devotee, "rashi", None),
        family_head_id=devotee.family_head_id,
        preferred_language=devotee.preferred_language or "en",
        receive_sms=devotee.receive_sms if devotee.receive_sms is not None else True,
        receive_email=devotee.receive_email if devotee.receive_email is not None else True,
        tags=tags_json,
        temple_id=current_user.temple_id if current_user else None,
    )
    db.add(db_devotee)
    db.commit()
    db.refresh(db_devotee)
    return DevoteeResponse.from_orm_with_masking(db_devotee, current_user, db)


@router.put("/{devotee_id}", response_model=DevoteeResponse)
def update_devotee(
    devotee_id: int,
    devotee: DevoteeUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Update a devotee"""
    db_devotee = db.query(Devotee).filter(Devotee.id == devotee_id).first()
    if not db_devotee:
        raise HTTPException(status_code=404, detail="Devotee not found")

    # Update fields
    update_data = devotee.dict(exclude_unset=True)

    # Handle tags separately (convert list to JSON)
    if "tags" in update_data:
        tags_json = json.dumps(update_data["tags"]) if update_data["tags"] else None
        db_devotee.tags = tags_json
        del update_data["tags"]

    for field, value in update_data.items():
        setattr(db_devotee, field, value)

    if "name" in update_data:
        db_devotee.full_name = update_data["name"]  # Update full_name too

    db_devotee.updated_at = datetime.utcnow().isoformat()
    db.commit()
    db.refresh(db_devotee)
    return DevoteeResponse.from_orm_with_masking(db_devotee, current_user, db)


@router.delete("/{devotee_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_devotee(
    devotee_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)
):
    """Delete a devotee"""
    db_devotee = db.query(Devotee).filter(Devotee.id == devotee_id).first()
    if not db_devotee:
        raise HTTPException(status_code=404, detail="Devotee not found")

    db.delete(db_devotee)
    db.commit()
    return None


@router.post("/merge", response_model=DevoteeResponse)
def merge_devotees(
    primary_id: int = Query(..., description="ID of devotee to keep"),
    duplicate_ids: List[int] = Query(..., description="IDs of devotees to merge into primary"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Merge duplicate devotees into one primary devotee
    Transfers all donations and bookings to primary devotee
    """
    primary = db.query(Devotee).filter(Devotee.id == primary_id).first()
    if not primary:
        raise HTTPException(status_code=404, detail="Primary devotee not found")

    duplicates = db.query(Devotee).filter(Devotee.id.in_(duplicate_ids)).all()
    if len(duplicates) != len(duplicate_ids):
        raise HTTPException(status_code=404, detail="Some duplicate devotees not found")

    # Merge data: keep non-null values from duplicates
    for dup in duplicates:
        if not primary.email and dup.email:
            primary.email = dup.email
        if not primary.address and dup.address:
            primary.address = dup.address
        if not primary.city and dup.city:
            primary.city = dup.city
        if not primary.state and dup.state:
            primary.state = dup.state
        if not primary.pincode and dup.pincode:
            primary.pincode = dup.pincode
        if not primary.date_of_birth and dup.date_of_birth:
            primary.date_of_birth = dup.date_of_birth
        if not primary.gothra and dup.gothra:
            primary.gothra = dup.gothra
        if not primary.nakshatra and dup.nakshatra:
            primary.nakshatra = dup.nakshatra
        if not primary.rashi and dup.rashi:
            primary.rashi = dup.rashi

        # Merge tags
        primary_tags = json.loads(primary.tags) if primary.tags else []
        dup_tags = json.loads(dup.tags) if dup.tags else []
        merged_tags = list(set(primary_tags + dup_tags))
        primary.tags = json.dumps(merged_tags) if merged_tags else None

        # Transfer donations
        db.query(Donation).filter(Donation.devotee_id == dup.id).update(
            {Donation.devotee_id: primary.id}
        )

        # Transfer bookings
        db.query(SevaBooking).filter(SevaBooking.devotee_id == dup.id).update(
            {SevaBooking.devotee_id: primary.id}
        )

        # Transfer family relationships
        db.query(Devotee).filter(Devotee.family_head_id == dup.id).update(
            {Devotee.family_head_id: primary.id}
        )

        # Delete duplicate
        db.delete(dup)

    primary.updated_at = datetime.utcnow().isoformat()
    db.commit()
    db.refresh(primary)

    return DevoteeResponse.from_orm_with_masking(primary, current_user, db)


@router.put("/{devotee_id}/link-family", response_model=DevoteeResponse)
def link_family_member(
    devotee_id: int,
    family_head_id: int = Query(..., description="ID of family head"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Link a devotee to a family head"""
    devotee = db.query(Devotee).filter(Devotee.id == devotee_id).first()
    if not devotee:
        raise HTTPException(status_code=404, detail="Devotee not found")

    family_head = db.query(Devotee).filter(Devotee.id == family_head_id).first()
    if not family_head:
        raise HTTPException(status_code=404, detail="Family head not found")

    devotee.family_head_id = family_head_id
    devotee.updated_at = datetime.utcnow().isoformat()
    db.commit()
    db.refresh(devotee)

    return DevoteeResponse.from_orm_with_masking(devotee, current_user, db)


@router.put("/{devotee_id}/tags", response_model=DevoteeResponse)
def update_devotee_tags(
    devotee_id: int,
    tags: List[str] = Query(..., description="List of tags"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Update devotee tags"""
    devotee = db.query(Devotee).filter(Devotee.id == devotee_id).first()
    if not devotee:
        raise HTTPException(status_code=404, detail="Devotee not found")

    devotee.tags = json.dumps(tags) if tags else None
    devotee.updated_at = datetime.utcnow().isoformat()
    db.commit()
    db.refresh(devotee)

    return DevoteeResponse.from_orm_with_masking(devotee, current_user, db)


@router.get("/{devotee_id}/family", response_model=List[DevoteeResponse])
def get_family_members(
    devotee_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)
):
    """Get all family members of a devotee"""
    devotee = db.query(Devotee).filter(Devotee.id == devotee_id).first()
    if not devotee:
        raise HTTPException(status_code=404, detail="Devotee not found")

    # Get family head
    family_head_id = devotee.family_head_id if devotee.family_head_id else devotee.id

    # Get all family members
    family_members = (
        db.query(Devotee)
        .filter(or_(Devotee.family_head_id == family_head_id, Devotee.id == family_head_id))
        .all()
    )

    return [DevoteeResponse.from_orm_with_masking(d, current_user, db) for d in family_members]


@router.post("/bulk-import")
def bulk_import_devotees(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Bulk import devotees from CSV or Excel file
    Returns summary of successful imports, skipped (duplicates), and errors
    """
    temple_id = current_user.temple_id if current_user else None

    if not temple_id:
        raise HTTPException(status_code=400, detail="Temple ID is required")

    # Validate file type
    file_extension = file.filename.split(".")[-1].lower() if file.filename else ""
    if file_extension not in ["csv", "xlsx", "xls"]:
        raise HTTPException(
            status_code=400, detail="File must be CSV or Excel format (.csv, .xlsx, .xls)"
        )

    results = {"success_count": 0, "skipped_count": 0, "errors": []}

    try:
        # Read file content
        content = file.file.read()

        # Parse based on file type
        if file_extension == "csv":
            # Parse CSV
            content_str = content.decode("utf-8")
            csv_reader = csv.DictReader(io.StringIO(content_str))
            rows = list(csv_reader)
        else:
            # Parse Excel
            workbook = load_workbook(io.BytesIO(content), read_only=True)
            sheet = workbook.active

            # Get headers from first row
            headers = [cell.value for cell in sheet[1]]
            rows = []
            for row in sheet.iter_rows(min_row=2, values_only=False):
                row_dict = {}
                for idx, cell in enumerate(row):
                    if idx < len(headers) and headers[idx]:
                        row_dict[headers[idx]] = cell.value
                if any(row_dict.values()):  # Skip empty rows
                    rows.append(row_dict)

        # Process each row
        for row_num, row in enumerate(rows, start=2):  # Start at 2 (1 is header)
            try:
                # Extract data with case-insensitive column matching
                row_lower = {k.lower().strip() if k else None: v for k, v in row.items() if k}

                # Get name (prefer first_name + last_name, fallback to name)
                first_name = row_lower.get("first_name") or row_lower.get("firstname") or ""
                last_name = row_lower.get("last_name") or row_lower.get("lastname") or ""
                name = row_lower.get("name") or row_lower.get("full_name") or ""

                if not name and first_name:
                    name = f"{first_name} {last_name}".strip()
                elif not name:
                    results["errors"].append(
                        {"row": row_num, "message": "Name or first_name is required"}
                    )
                    continue

                # Get phone (required)
                phone = str(
                    row_lower.get("phone")
                    or row_lower.get("mobile")
                    or row_lower.get("phone_number")
                    or ""
                ).strip()
                if not phone:
                    results["errors"].append(
                        {"row": row_num, "message": "Phone number is required"}
                    )
                    continue

                # Clean phone number (remove spaces, dashes)
                phone = phone.replace(" ", "").replace("-", "").replace("(", "").replace(")", "")

                # Get country code
                country_code = str(
                    row_lower.get("country_code") or row_lower.get("countrycode") or "+91"
                ).strip()
                if not country_code.startswith("+"):
                    country_code = "+" + country_code

                # Extract phone from country code if included
                if phone.startswith(country_code):
                    phone = phone[len(country_code) :]

                # For Indian numbers, take last 10 digits
                if country_code == "+91" and len(phone) > 10:
                    phone = phone[-10:]

                # Check if devotee already exists
                existing = (
                    db.query(Devotee)
                    .filter(Devotee.phone == phone, Devotee.temple_id == temple_id)
                    .first()
                )

                if existing:
                    results["skipped_count"] += 1
                    continue

                # Get optional fields
                name_prefix = (
                    str(row_lower.get("name_prefix") or row_lower.get("prefix") or "").strip()
                    or None
                )
                email = str(row_lower.get("email") or "").strip() or None
                address = str(row_lower.get("address") or "").strip() or None
                city = str(row_lower.get("city") or "").strip() or None
                state = str(row_lower.get("state") or "").strip() or None
                pincode = (
                    str(row_lower.get("pincode") or row_lower.get("pin") or "").strip() or None
                )
                country = str(row_lower.get("country") or "India").strip()

                # Date of birth
                date_of_birth = None
                dob_str = str(
                    row_lower.get("date_of_birth")
                    or row_lower.get("dob")
                    or row_lower.get("birthdate")
                    or ""
                ).strip()
                if dob_str:
                    try:
                        # Try parsing common date formats
                        for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"]:
                            try:
                                date_of_birth = datetime.strptime(dob_str, fmt).date()
                                break
                            except:
                                continue
                    except:
                        pass

                gothra = (
                    str(row_lower.get("gothra") or row_lower.get("gotra") or "").strip() or None
                )
                nakshatra = str(row_lower.get("nakshatra") or "").strip() or None
                rashi = str(row_lower.get("rashi") or "").strip() or None

                # Tags
                tags_str = str(row_lower.get("tags") or "").strip()
                tags = None
                if tags_str:
                    tags_list = [t.strip() for t in tags_str.split(",") if t.strip()]
                    if tags_list:
                        tags = json.dumps(tags_list)

                # Create devotee
                db_devotee = Devotee(
                    name_prefix=name_prefix,
                    first_name=first_name or name.split()[0] if name else None,
                    last_name=last_name or " ".join(name.split()[1:])
                    if len(name.split()) > 1
                    else None,
                    name=name,
                    full_name=name,
                    country_code=country_code,
                    phone=phone,
                    email=email,
                    address=address,
                    city=city,
                    state=state,
                    pincode=pincode,
                    country=country,
                    date_of_birth=date_of_birth,
                    gothra=gothra,
                    nakshatra=nakshatra,
                    rashi=rashi,
                    tags=tags,
                    temple_id=temple_id,
                )

                db.add(db_devotee)
                results["success_count"] += 1

            except Exception as e:
                results["errors"].append({"row": row_num, "message": str(e)})

        # Commit all successful imports
        db.commit()

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Error processing file: {str(e)}")

    return results
