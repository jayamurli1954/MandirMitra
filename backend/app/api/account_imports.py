import csv
import io
from pathlib import Path

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.api.accounts import _resolve_temple_id, _seed_default_accounts_for_temple
from app.core.database import get_db
from app.core.role_permissions import require_action_permission
from app.core.security import get_current_user
from app.core.temple_context import require_temple_write_access_to_temple
from app.models.accounting import Account, AccountSubType, AccountType
from app.models.user import User
from app.schemas.accounting import AccountCreate

router = APIRouter(prefix="/api/v1/accounts", tags=["accounts"])

LEGACY_ACCOUNT_CODE_MAP = {
    "5101": "52003",
    "5102": "52001",
    "5110": "53002",
    "5111": "53007",
    "5120": "53004",
    "5201": "54006",
    "5202": "51004",
    "5203": "54007",
    "5301": "51001",
    "5302": "54012",
    "5401": "54005",
    "5402": "54005",
    "5403": "54004",
    "5501": "54010",
    "5502": "54001",
    "5503": "53006",
}

CODE_TO_TYPE_MAP = {
    "1": AccountType.ASSET,
    "2": AccountType.LIABILITY,
    "3": AccountType.EQUITY,
    "4": AccountType.INCOME,
    "5": AccountType.EXPENSE,
}


def _require_account_admin(current_user: User) -> None:
    if current_user.role not in ["admin", "temple_manager", "accountant", "super_admin"]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only admin users can manage accounts",
        )


def _normalize_account_code(raw_code: str) -> str:
    normalized = str(raw_code or "").strip()
    if not normalized:
        raise ValueError("account_code or legacy_code is required")
    if normalized in LEGACY_ACCOUNT_CODE_MAP:
        return LEGACY_ACCOUNT_CODE_MAP[normalized]
    if normalized.isdigit() and len(normalized) == 4:
        return f"{normalized[:3]}0{normalized[3:]}"
    return normalized


def _parse_import_rows(file_name: str, file_bytes: bytes) -> list[dict]:
    suffix = Path(file_name or "").suffix.lower()
    if suffix == ".csv":
        return list(csv.DictReader(io.StringIO(file_bytes.decode("utf-8-sig"))))
    if suffix in {".xlsx", ".xlsm"}:
        workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(value).strip() if value is not None else "" for value in rows[0]]
        parsed_rows = []
        for row in rows[1:]:
            if not any(cell not in (None, "") for cell in row):
                continue
            parsed_rows.append(
                {
                    headers[index]: row[index]
                    for index in range(min(len(headers), len(row)))
                    if headers[index]
                }
            )
        return parsed_rows
    raise HTTPException(status_code=400, detail="Unsupported file format. Use .csv or .xlsx")


def _coerce_account_type(raw_type: str | None, normalized_code: str) -> AccountType:
    if raw_type:
        return AccountType(str(raw_type).strip().lower())
    if not normalized_code.isdigit() or len(normalized_code) != 5:
        raise ValueError("account_type is required when code is not a valid 5-digit code")
    return CODE_TO_TYPE_MAP[normalized_code[0]]


def _coerce_account_subtype(raw_subtype: str | None) -> AccountSubType | None:
    normalized = str(raw_subtype or "").strip().lower()
    return AccountSubType(normalized) if normalized else None


def _derive_opening_balances(account_type: AccountType, row: dict) -> tuple[float, float]:
    debit = float(row.get("opening_balance_debit") or 0 or 0.0)
    credit = float(row.get("opening_balance_credit") or 0 or 0.0)
    signed_balance = row.get("opening_balance")
    if debit or credit or signed_balance in (None, ""):
        return max(0.0, debit), max(0.0, credit)

    numeric_balance = float(signed_balance)
    if account_type in (AccountType.ASSET, AccountType.EXPENSE):
        return max(0.0, numeric_balance), max(0.0, abs(numeric_balance) if numeric_balance < 0 else 0.0)
    return max(0.0, abs(numeric_balance) if numeric_balance < 0 else 0.0), max(0.0, numeric_balance)


@router.post("/import-legacy", response_model=dict)
async def import_legacy_accounts(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Import legacy account masters from CSV/XLSX into the current COA."""
    _require_account_admin(current_user)
    temple_id = _resolve_temple_id(db, current_user)
    require_temple_write_access_to_temple(db, current_user, temple_id)
    try:
        require_action_permission(
            db,
            current_user,
            "import_legacy_accounts",
            temple_id=temple_id,
            detail="You do not have permission to import legacy accounts",
        )
    except PermissionError as exc:
        raise HTTPException(status_code=403, detail=str(exc)) from exc
    _seed_default_accounts_for_temple(db, temple_id)

    if not file.filename:
        raise HTTPException(status_code=400, detail="File name is required")

    rows = _parse_import_rows(file.filename, await file.read())
    if not rows:
        raise HTTPException(status_code=400, detail="Import file is empty")

    accounts_by_code = {
        account.account_code: account
        for account in db.query(Account).filter(Account.temple_id == temple_id).all()
    }

    created = []
    updated = []
    errors = []

    for row_number, row in enumerate(rows, start=2):
        try:
            normalized_code = _normalize_account_code(
                row.get("account_code") or row.get("legacy_code") or row.get("code") or ""
            )
            account_name = str(row.get("account_name") or row.get("name") or "").strip()
            if not account_name:
                raise ValueError("account_name is required")

            account_type = _coerce_account_type(row.get("account_type"), normalized_code)
            account_subtype = _coerce_account_subtype(row.get("account_subtype"))
            description = str(row.get("description") or "").strip() or None
            opening_debit, opening_credit = _derive_opening_balances(account_type, row)

            existing_account = accounts_by_code.get(normalized_code)
            if existing_account:
                existing_account.account_name = account_name
                existing_account.description = description
                if account_subtype is not None:
                    existing_account.account_subtype = account_subtype
                existing_account.opening_balance_debit = opening_debit
                existing_account.opening_balance_credit = opening_credit
                updated.append(normalized_code)
                continue

            payload = AccountCreate(
                account_code=normalized_code,
                account_name=account_name,
                description=description,
                account_type=account_type,
                account_subtype=account_subtype,
                temple_id=temple_id,
                opening_balance_debit=opening_debit,
                opening_balance_credit=opening_credit,
                is_active=True,
                allow_manual_entry=True,
            )
            account = Account(**payload.dict())
            db.add(account)
            db.flush()
            accounts_by_code[normalized_code] = account
            created.append(normalized_code)
        except Exception as exc:
            errors.append({"row": row_number, "error": str(exc)})

    db.commit()
    return {
        "message": f"Processed {len(rows)} legacy account rows",
        "created_count": len(created),
        "updated_count": len(updated),
        "created": created,
        "updated": updated,
        "errors": errors,
    }
