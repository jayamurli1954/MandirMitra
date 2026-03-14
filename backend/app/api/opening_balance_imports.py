import csv
import io
from pathlib import Path

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.role_permissions import require_action_permission
from app.core.security import get_current_user
from app.core.temple_context import require_temple_write_access
from app.models.accounting import Account, AccountType
from app.models.user import User

router = APIRouter(prefix="/api/v1/opening-balances", tags=["opening-balances"])

LEGACY_ACCOUNT_CODE_MAP = {
    "5101": "52003",
    "5102": "52001",
    "5110": "53002",
    "5111": "53007",
    "5120": "53004",
    "5201": "54006",
    "5202": "51004",
    "5203": "54007",
    "5301": "51001",
    "5302": "54012",
    "5401": "54005",
    "5402": "54005",
    "5403": "54004",
    "5501": "54010",
    "5502": "54001",
    "5503": "53006",
}


def _require_admin_access(current_user: User) -> None:
    if current_user.role not in ["admin", "temple_manager", "accountant", "super_admin"]:
        raise HTTPException(status_code=403, detail="Only admin users can manage opening balances")


def _normalize_account_code(raw_code: str) -> str:
    normalized = str(raw_code or "").strip()
    if not normalized:
        raise ValueError("account_code or legacy_code is required")
    if normalized in LEGACY_ACCOUNT_CODE_MAP:
        return LEGACY_ACCOUNT_CODE_MAP[normalized]
    if normalized.isdigit() and len(normalized) == 4:
        return f"{normalized[:3]}0{normalized[3:]}"
    return normalized


def _parse_import_rows(file_name: str, file_bytes: bytes) -> list[dict]:
    suffix = Path(file_name or "").suffix.lower()
    if suffix == ".csv":
        return list(csv.DictReader(io.StringIO(file_bytes.decode("utf-8-sig"))))
    if suffix in {".xlsx", ".xlsm"}:
        workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(value).strip() if value is not None else "" for value in rows[0]]
        parsed_rows = []
        for row in rows[1:]:
            if not any(cell not in (None, "") for cell in row):
                continue
            parsed_rows.append(
                {
                    headers[index]: row[index]
                    for index in range(min(len(headers), len(row)))
                    if headers[index]
                }
            )
        return parsed_rows
    raise HTTPException(status_code=400, detail="Unsupported file format. Use .csv or .xlsx")


@router.post("/import")
async def import_opening_balances(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Import opening balances from CSV/XLSX for balance-sheet accounts."""
    _require_admin_access(current_user)
    temple_id = require_temple_write_access(db, current_user, active_only=False)
    try:
        require_action_permission(
            db,
            current_user,
            "manage_opening_balances",
            temple_id=temple_id,
            detail="You do not have permission to manage opening balances",
        )
    except PermissionError as exc:
        raise HTTPException(status_code=403, detail=str(exc)) from exc

    if not file.filename:
        raise HTTPException(status_code=400, detail="File name is required")

    rows = _parse_import_rows(file.filename, await file.read())
    if not rows:
        raise HTTPException(status_code=400, detail="Import file is empty")

    accounts_by_code = {
        account.account_code: account
        for account in db.query(Account).filter(Account.temple_id == temple_id).all()
    }

    updated = []
    errors = []
    for row_number, row in enumerate(rows, start=2):
        try:
            normalized_code = _normalize_account_code(
                row.get("account_code") or row.get("legacy_code") or row.get("code") or ""
            )
            account = accounts_by_code.get(normalized_code)
            if not account:
                raise ValueError(f"Account '{normalized_code}' not found")
            if account.account_type not in [AccountType.ASSET, AccountType.LIABILITY, AccountType.EQUITY]:
                raise ValueError("Only balance sheet accounts can have opening balances")

            debit = row.get("opening_balance_debit")
            credit = row.get("opening_balance_credit")
            signed_balance = row.get("opening_balance")

            if debit not in (None, ""):
                account.opening_balance_debit = max(0.0, float(debit))
            if credit not in (None, ""):
                account.opening_balance_credit = max(0.0, float(credit))

            if debit in (None, "") and credit in (None, "") and signed_balance not in (None, ""):
                numeric_balance = float(signed_balance)
                if account.account_type == AccountType.ASSET:
                    account.opening_balance_debit = max(0.0, numeric_balance)
                    account.opening_balance_credit = max(0.0, abs(numeric_balance) if numeric_balance < 0 else 0.0)
                else:
                    account.opening_balance_credit = max(0.0, numeric_balance)
                    account.opening_balance_debit = max(0.0, abs(numeric_balance) if numeric_balance < 0 else 0.0)

            updated.append(
                {
                    "account_code": account.account_code,
                    "account_name": account.account_name,
                    "opening_balance_debit": account.opening_balance_debit,
                    "opening_balance_credit": account.opening_balance_credit,
                }
            )
        except Exception as exc:
            errors.append({"row": row_number, "error": str(exc)})

    db.commit()
    return {
        "message": f"Processed {len(rows)} opening balance rows",
        "updated_count": len(updated),
        "updated": updated,
        "errors": errors,
    }
